#!/usr/bin/env python3
"""Publish Week 1 Rank 3: Heart-Touching Love Proposal Quotes 2027."""
import base64
import json
import mimetypes
import os
import re
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from html import escape

ROOT = Path(__file__).resolve().parents[1]
USER = os.environ["WP_USER"]
PWD = os.environ["WP_APP_PASSWORD"]
TOKEN = base64.b64encode(f"{USER}:{PWD}".encode()).decode()
AUTH = {"Authorization": f"Basic {TOKEN}", "User-Agent": "BluestoneSEO/1.0"}
API = "https://blog.bluestone.com/wp-json/wp/v2"

TITLE = "75+ Heart-Touching Love Proposal Quotes & Propose Day Wishes for 2027"
SLUG = "heart-touching-love-proposal-quotes-2027"
META_DESC = (
    "Copy-ready love proposal quotes, propose day wishes and proposal lines for her or him for 2027. "
    "Short WhatsApp lines, romantic messages and captions to help you say yes."
)
FOCUS_KW = "propose a girl"
YOAST_TITLE = "Love Proposal Quotes & Propose Day Wishes 2027 | BlueStone"

PRODUCTS = [
    {
        "name": "The Anya Ring",
        "url": "https://www.bluestone.com/rings/the-anya-ring~7515.html",
        "png": ROOT / "ProductImages/seo images/Rings/The Anya Ring.png",
    },
    {
        "name": "The Liza Ring",
        "url": "https://www.bluestone.com/rings/the-liza-ring~7623.html",
        "png": ROOT / "ProductImages/seo images/Rings/The Liza ring.png",
    },
    {
        "name": "The Quinn Ring",
        "url": "https://www.bluestone.com/rings/the-quinn-ring~57845.html",
        "png": ROOT / "ProductImages/seo images/Rings/The Quinn Ring.png",
    },
    {
        "name": "The Rafia Ring",
        "url": "https://www.bluestone.com/rings/the-rafia-ring~53638.html",
        "png": ROOT / "ProductImages/seo images/Rings/The Rafia Ring.png",
    },
    {
        "name": "The Gigi Ring",
        "url": "https://www.bluestone.com/rings/the-gigi-ring~64382.html",
        "png": ROOT / "ProductImages/seo images/Rings/The Gigi Ring.png",
    },
    {
        "name": "The Ebony Ring",
        "url": "https://www.bluestone.com/rings/the-ebony-ring~9686.html",
        "png": ROOT / "ProductImages/seo images/Rings/The Ebony Ring.png",
    },
]

TYPE3 = [
    {
        "key": "hero",
        "src": Path("/Users/vikasindoria/.cursor/projects/Users-vikasindoria-Documents-Geo-and-Seo/assets/proposal-hero-mood-2027.png"),
        "filename": "proposal-hero-mood-2027.webp",
        "alt": "Heart-touching love proposal mood for Propose Day 2027, ring box and romantic setting",
        "featured": True,
    },
    {
        "key": "flatlay",
        "src": Path("/Users/vikasindoria/.cursor/projects/Users-vikasindoria-Documents-Geo-and-Seo/assets/proposal-flatlay-phone-2027.png"),
        "filename": "proposal-flatlay-phone-2027.webp",
        "alt": "Short love proposal message draft on phone with ring box for Propose Day 2027",
        "featured": False,
    },
    {
        "key": "lifestyle",
        "src": Path("/Users/vikasindoria/.cursor/projects/Users-vikasindoria-Documents-Geo-and-Seo/assets/proposal-lifestyle-couple-2027.png"),
        "filename": "proposal-lifestyle-couple-2027.webp",
        "alt": "Couple proposal moment, love proposal quotes inspiration for Propose Day 2027",
        "featured": False,
    },
]

SECTIONS = {
    "whatsapp": [
        "Will you make my every ordinary day feel chosen? Say yes.",
        "I do not need perfect words. I need you. Marry me?",
        "You are my calm and my favourite adventure. Be mine.",
        "Life makes more sense with you in it. Propose accepted?",
        "I want a future that sounds like us laughing. Always?",
        "You are the person I want to text first and last. Forever?",
        "I choose you today, tomorrow, and every after that.",
        "Can we turn this love into a lifetime plan?",
        "You make commitment feel easy. Will you stay?",
        "My heart already said yes. I hope yours does too.",
    ],
    "quotes": [
        "You walked into my life quietly and changed everything.",
        "Loving you taught me what choosing someone every day really means.",
        "Every version of my future includes your smile.",
        "You turned small moments into memories I never want to lose.",
        "If love has a language, mine speaks your name.",
        "I am better because you believe in me.",
        "With you, forever feels less like a promise and more like a plan.",
        "You are my certainty in a changing world.",
        "I want a life where we grow, not just glow.",
        "Some people feel like home without trying. You are that person.",
    ],
    "propose_day": [
        "Happy Propose Day 2027! Today I choose you, honestly and fully.",
        "On Propose Day, I want to begin our story with a clear yes.",
        "This Propose Day feels special because of us.",
        "Propose Day reminder: I still choose you.",
        "February brings roses. I bring a lifetime promise.",
        "Propose Day 2027: may our love stay steady and playful.",
        "Today feels like the right moment to ask for your forever.",
        "Happy Propose Day! You are my favourite decision.",
        "Let us turn this feeling into a future we build together.",
        "Propose Day wishes to the one who makes love feel real.",
    ],
    "her": [
        "I admire your strength and your kindness. Will you choose me?",
        "You inspire me to be better. Stay with me.",
        "I want a future shaped by us, not just by me.",
        "You make love feel honest and safe.",
        "I do not want perfection. I want you.",
        "Let us build something real together.",
        "You are my calm and my excitement.",
        "I want to walk forward with you.",
        "Every plan I make has you in the centre.",
        "You are the love I want to grow with.",
    ],
    "boy": [
        "I know what I want, and it is you.",
        "I choose you with confidence and love.",
        "You feel like my future.",
        "Life feels aligned when I imagine us.",
        "I want to grow old laughing with you.",
        "This is me asking for our forever.",
        "Will you build a life with me?",
        "I do not wait for signs. I choose you.",
        "Being with you feels like progress.",
        "Today I ask for your future, not just your heart.",
    ],
    "girl_propose_boy": [
        "I am asking first because my heart is sure.",
        "You make me brave enough to say this out loud.",
        "I want us, openly and completely.",
        "Will you let me love you out loud?",
        "I choose you without waiting for the perfect scene.",
        "My love is not shy about wanting you.",
        "If I am proposing, it is because you are worth it.",
        "I want a partnership that feels like us.",
        "Say yes to a future we write together.",
        "I am here, certain, and asking.",
    ],
    "chat_boy": [
        "Can I tell you something serious for a second? I want us long term.",
        "Random thought: I see my future with you in it.",
        "Not a meme, just truth: I want to build something with you.",
        "If we keep going like this, I do not want it to stay casual.",
        "You make everyday chats feel like home.",
        "I think we are more than good chemistry. Want to find out?",
        "Would you be open to making this official?",
        "I like you in a way that deserves a real conversation.",
        "Can we talk about us, properly?",
        "I am ready for the next chapter if you are.",
    ],
    "shy": [
        "I love you, and I want to spend my life with you.",
        "Life feels better with you. Will you stay?",
        "You are my person. Marry me?",
        "I am happiest when I am with you. Always?",
        "I do not need big words. Just you.",
        "Let us choose each other every day.",
        "This feels right because it is us.",
        "I want you by my side for life.",
        "Some promises begin quietly.",
        "I already think in terms of us.",
    ],
    "captions": [
        "Choosing you feels natural.",
        "Propose Day, honest hearts.",
        "Love feels steady with you.",
        "A quiet yes to us.",
        "This week, I choose love.",
        "Forever starts with one brave sentence.",
        "My favourite plot twist: us.",
        "Heart full, question ready.",
        "Proposal energy only.",
        "Say yes to ordinary magic.",
    ],
}


def api(method, path, data=None, raw_body=None, headers=None):
    h = dict(AUTH)
    if headers:
        h.update(headers)
    if data is not None:
        h["Content-Type"] = "application/json"
        body = json.dumps(data).encode()
    else:
        body = raw_body
    req = urllib.request.Request(f"{API}/{path}", data=body, headers=h, method=method)
    with urllib.request.urlopen(req, timeout=180) as r:
        return json.loads(r.read().decode())


def to_webp(src: Path, dest: Path, size=None, carousel=False):
    from PIL import Image

    im = Image.open(src).convert("RGB")
    if carousel:
        tw, th = 960, 535
        im.thumbnail((tw, th), Image.Resampling.LANCZOS)
        canvas = Image.new("RGB", (tw, th), (245, 243, 240))
        canvas.paste(im, ((tw - im.width) // 2, (th - im.height) // 2))
        im = canvas
    elif size:
        im.thumbnail(size, Image.Resampling.LANCZOS)
    dest.parent.mkdir(parents=True, exist_ok=True)
    im.save(dest, "WEBP", quality=82, method=6)
    return im.size


def upload_media(path: Path, alt: str):
    data = path.read_bytes()
    ctype = "image/webp"
    headers = {
        "Content-Disposition": f'attachment; filename="{path.name}"',
        "Content-Type": ctype,
    }
    media = api("POST", "media", raw_body=data, headers=headers)
    api("POST", f"media/{media['id']}", {"alt_text": alt, "title": path.stem})
    return media


def ol(items):
    lines = ['<!-- wp:list {"ordered":true} -->', '<ol class="wp-block-list">']
    for item in items:
        lines.append(f"<li>{escape(item)}</li>")
    lines += ["</ol>", "<!-- /wp:list -->"]
    return "\n".join(lines)


def h2(text):
    return f'<!-- wp:heading -->\n<h2 class="wp-block-heading">{escape(text)}</h2>\n<!-- /wp:heading -->'


def h3(text):
    return f'<!-- wp:heading {{"level":3}} -->\n<h3 class="wp-block-heading">{escape(text)}</h3>\n<!-- /wp:heading -->'


def para(text):
    return f'<!-- wp:paragraph -->\n<p>{text}</p>\n<!-- /wp:paragraph -->'


def image_block(mid, src, alt, w, h):
    return (
        f'<!-- wp:image {{"id":{mid},"sizeSlug":"large","linkDestination":"none"}} -->\n'
        f'<figure class="wp-block-image size-large">'
        f'<img src="{src}" alt="{escape(alt)}" class="wp-image-{mid}" width="{w}" height="{h}" loading="lazy" decoding="async"/>'
        f"</figure>\n<!-- /wp:image -->"
    )


def build_carousel(product_media):
    cards = []
    dots = []
    for i, p in enumerate(product_media):
        alt = f"Love proposal gift idea: {p['name']} from BlueStone"
        cards.append(
            f'    <div class="bs-cf-card" data-i="{i}">\n'
            f'      <a class="bs-cf-media" href="{p["url"]}">\n'
            f'        <img src="{p["src"]}" alt="{escape(alt)}" width="960" height="535" loading="lazy" decoding="async"/>\n'
            f"      </a>\n"
            f'      <div class="bs-cf-meta">\n'
            f'        <p class="bs-cf-name">{escape(p["name"])}</p>\n'
            f'        <a class="bs-cf-cta" href="{p["url"]}">Buy now</a>\n'
            f"      </div>\n"
            f"    </div>"
        )
        active = " is-active" if i == 0 else ""
        dots.append(
            f'    <button type="button" class="bs-cf-dot{active}" data-i="{i}" aria-label="Product {i+1}"></button>'
        )
    style = Path(ROOT / "output/_eid_carousel_6_snippet.html").read_text().split("<style>")[1].split("</style>")[0]
    script = Path(ROOT / "output/_eid_carousel_6_snippet.html").read_text().split("<script>")[1].split("</script>")[0]
    script = script.replace("bs-cf-eid", "bs-cf-propose")
    return (
        "<!-- wp:html -->\n<style>\n"
        + style
        + '\n</style>\n<div class="bs-cf" id="bs-cf-propose" data-interval="3200" aria-roledescription="carousel" aria-label="BlueStone proposal ring gift ideas">\n'
        + '  <button type="button" class="bs-cf-nav bs-cf-prev" aria-label="Previous">&#8249;</button>\n'
        + '  <button type="button" class="bs-cf-nav bs-cf-next" aria-label="Next">&#8250;</button>\n'
        + '  <div class="bs-cf-stage">\n'
        + "\n".join(cards)
        + "\n  </div>\n  <div class=\"bs-cf-dots\" role=\"tablist\">\n"
        + "\n".join(dots)
        + "\n  </div>\n</div>\n<script>\n"
        + script
        + "\n</script>\n<!-- /wp:html -->"
    )


def build_faqs():
    faqs = [
        (
            "What are some short love proposal quotes for WhatsApp?",
            "Try: \"You are my person. Marry me?\" or \"Life feels better with you. Will you stay?\" Short proposal lines work best when they sound like you, not like a greeting card.",
        ),
        (
            "What should I say when I propose to a girl in 2027?",
            "Keep it honest and specific. Name one thing you love about her, one future you want together, then ask clearly. Proposal lines for her work best when they feel personal, not performative.",
        ),
        (
            "What are good Propose Day wishes for my partner?",
            "Happy Propose Day 2027 wishes can be warm and direct: \"Today I choose you, honestly and fully\" or \"You are my favourite decision.\" Match the tone to your relationship stage.",
        ),
        (
            "How can a girl propose to a boy?",
            "Lead with clarity, not drama. Romantic girl propose to boy lines can be playful or confident: \"I know what I want, and it is you.\" Ask in a setting where he can respond honestly.",
        ),
        (
            "What are simple proposal quotes for shy people?",
            "Use one sentence: \"I love you, and I want to spend my life with you.\" Shy proposal quotes do not need poetry. Eye contact and sincerity matter more than length.",
        ),
        (
            "What are the best proposal captions for Instagram?",
            "Keep captions under two lines: \"Choosing you feels natural\" or \"Forever starts with one brave sentence.\" Pair them with a photo that matches the mood, not just the ring.",
        ),
        (
            "How do you propose to a boy on chat?",
            "Move from banter to clarity: \"Can we talk about us, properly?\" or \"I see my future with you in it.\" Give him space to reply without pressure.",
        ),
    ]
    html = [h2("Frequently Asked Questions about Love Proposal Quotes")]
    schema = []
    for q, a in faqs:
        html.append(h3(q))
        html.append(para(escape(a)))
        schema.append(
            {
                "@type": "Question",
                "name": q,
                "acceptedAnswer": {"@type": "Answer", "text": a},
            }
        )
    return "\n\n".join(html), schema


def build_content(flatlay, lifestyle, carousel):
    faq_html, faq_schema = build_faqs()
    parts = [
        '<!-- wp:html -->\n<style>\n.bs-eeat{margin:0 auto 1.25rem;max-width:720px;text-align:center;font-size:.95rem;color:#444;line-height:1.5}\n.bs-eeat strong{color:#111}\n.entry-content img,.wp-block-image img{max-width:100%;height:auto}\n</style>\n<!-- /wp:html -->',
        '<!-- wp:paragraph {"align":"center"} -->\n<p class="has-text-align-center bs-eeat">By <strong>Vikas</strong>, BlueStone Editorial</p>\n<!-- /wp:paragraph -->',
        para(
            "Looking for love proposal quotes you can copy, edit, and send today? "
            "Here are heart-touching proposal lines, Propose Day wishes, and short messages for WhatsApp, "
            "Instagram, and that one honest conversation, all ready for 2027."
        ),
        para(
            "<strong>TL;DR:</strong> Pick a one-line WhatsApp proposal quote for a quick send, "
            "a heartfelt paragraph for a letter or speech, and a short caption for your photo. "
            "Scroll the lists below and use what sounds like you."
        ),
        para(
            "A proposal is remembered less for the setting and more for the sentence you choose in that pause before the answer. "
            "The ring carries meaning, but the words give it a voice. These proposal quotes are grouped so you can find a line that fits your relationship, "
            "whether you are planning Propose Day 2027, a quiet chat, or a full \"will you marry me\" moment."
        ),
        h2("Short Love Proposal Messages for WhatsApp"),
        para("Use these when you need proposal lines for her or him that are warm, clear, and easy to forward."),
        ol(SECTIONS["whatsapp"]),
        image_block(flatlay["id"], flatlay["src"], flatlay["alt"], flatlay["w"], flatlay["h"]),
        h2("Heart-Touching Love Proposal Quotes in English"),
        para("These love proposal quotes suit cards, speeches, or a letter you want them to keep."),
        ol(SECTIONS["quotes"]),
        h2("Propose Day Wishes for 2027"),
        para("Propose Day falls in Valentine Week. These happy Propose Day wishes fit messages, cards, and status updates."),
        ol(SECTIONS["propose_day"]),
        h2("Romantic Proposal Lines to Propose to a Girl"),
        para("When your search is literally <em>propose a girl</em>, start with lines that feel admiring, steady, and specific."),
        ol(SECTIONS["her"]),
        h2("A soft proposal gift idea (if you are pairing words with a ring)"),
        para(
            "If you are gifting too, keep the focus on the moment first. "
            "A ring that feels like everyday you, not just occasion you, often lands better after the words. "
            "Browse these BlueStone rings if you want a starting point."
        ),
        carousel,
        h2("Proposal Lines to Propose to a Boy"),
        para("Modern love includes confident, direct proposal quotes for him. These lines keep the ask clear."),
        ol(SECTIONS["boy"]),
        image_block(lifestyle["id"], lifestyle["src"], lifestyle["alt"], lifestyle["w"], lifestyle["h"]),
        h2("Romantic Girl Propose to Boy Lines"),
        para("If you are searching <em>romantic girl propose boy</em>, these messages lead with certainty and warmth."),
        ol(SECTIONS["girl_propose_boy"]),
        h2("How to Propose to a Boy on Chat"),
        para("These proposal chat openers help you move from jokes to intent without ambush."),
        ol(SECTIONS["chat_boy"]),
        h2("Simple & Shy Proposal Quotes"),
        para("Not everyone wants a stage. These indirect and shy proposal quotes keep the pressure low."),
        ol(SECTIONS["shy"]),
        h2("Propose Day Captions for Instagram & Status"),
        para("Short captions for Propose Day posts, stories, and WhatsApp status."),
        ol(SECTIONS["captions"]),
        h2("How to pick the right proposal line in 10 seconds"),
        para("Match tone to setting: public asks need clear language; private moments allow softer lines."),
        para("Match length to medium: WhatsApp favours one sentence; letters can breathe."),
        para("Say it in your voice. If you would not normally use a word, cut it."),
        para("Practice once, then trust the moment. Perfect recall matters less than sincerity."),
        h2("More love & festive reads to explore"),
        para(
            'Planning other moments too? Save our '
            '<a href="https://blog.bluestone.com/happy-holi-wishes-messages-quotes-2027/">Happy Holi wishes for 2027</a>, '
            '<a href="https://blog.bluestone.com/happy-eid-mubarak-wishes-messages-quotes-2027/">Eid Mubarak wishes for 2027</a>, '
            '<a href="https://blog.bluestone.com/heart-touching-raksha-bandhan-wishes-quotes/">Raksha Bandhan wishes</a>, and '
            '<a href="https://blog.bluestone.com/romantic-valentines-day-wishes-quotes-celebrate-love-in-the-most-beautiful-way/">Valentine\'s Day wishes</a> '
            'for quick copy-ready messages across the year.'
        ),
        para(
            'For context on Propose Day and Valentine Week in India, see '
            '<a href="https://en.wikipedia.org/wiki/Propose_Day" rel="noopener">Propose Day on Wikipedia</a>.'
        ),
        h2("Conclusion"),
        para(
            "The best love proposal quotes sound like you on a brave day. Choose words that match your relationship, "
            "send them through the medium that feels natural, and let the moment breathe. "
            "Whether you are drafting Propose Day wishes for 2027 or asking for forever, sincerity beats performance every time."
        ),
        faq_html,
    ]
    return "\n\n".join(parts), faq_schema


def update_xlsx(post_id, url):
    import openpyxl

    wb = openpyxl.load_workbook(ROOT / "SEO Strategy 2026.xlsx")
    ws = wb["Week 1-2"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    for row in ws.iter_rows(min_row=2):
        if row[idx["Rank"]].value == 3:
            row[idx["Suggested URL Slug"]].value = SLUG
            row[idx["Bluestone Blog URL"]].value = url
            note = row[idx["Execution Note"]].value or ""
            row[idx["Execution Note"]].value = (
                f"{note.split('| Published')[0].strip()} | Published {datetime.now().strftime('%Y-%m-%d')} WP#{post_id} | "
                f"Refreshed Propose 2027; author Vikas; WebP hero + flatlay + lifestyle; 6-ring coverflow carousel"
            )
            break
    wb.save(ROOT / "SEO Strategy 2026.xlsx")


def main():
    assets = ROOT / "output/Week1_Rank3_Propose_assets"
    assets.mkdir(exist_ok=True)

    product_media = []
    for p in PRODUCTS:
        webp = assets / (p["name"].replace(" ", "-") + "-carousel.webp")
        to_webp(p["png"], webp, carousel=True)
        alt = f"Love proposal gift idea: {p['name']} from BlueStone"
        media = upload_media(webp, alt)
        product_media.append(
            {
                "name": p["name"],
                "url": p["url"],
                "id": media["id"],
                "src": media["source_url"],
            }
        )
        print("product", p["name"], media["id"])

    concept = {}
    hero_media = None
    for t in TYPE3:
        webp = assets / t["filename"]
        w, h = to_webp(t["src"], webp, size=(1400, 933))
        media = upload_media(webp, t["alt"])
        info = {"id": media["id"], "src": media["source_url"], "alt": t["alt"], "w": w, "h": h}
        concept[t["key"]] = info
        if t["featured"]:
            hero_media = media["id"]
        print("type3", t["key"], media["id"])

    carousel = build_carousel(product_media)
    content, faq_schema = build_content(concept["flatlay"], concept["lifestyle"], carousel)

    images = [concept["hero"]["src"], concept["flatlay"]["src"], concept["lifestyle"]["src"]] + [p["src"] for p in product_media]
    article_schema = {
        "@context": "https://schema.org",
        "@type": "BlogPosting",
        "headline": TITLE,
        "description": META_DESC,
        "datePublished": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "dateModified": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "author": {"@type": "Person", "name": "Vikas"},
        "publisher": {"@type": "Organization", "name": "BlueStone", "url": "https://www.bluestone.com/"},
        "image": images,
        "mainEntityOfPage": {
            "@type": "WebPage",
            "@id": f"https://blog.bluestone.com/{SLUG}/",
        },
        "keywords": [
            "propose a girl",
            "love proposal quotes",
            "propose day wishes",
            "proposal lines for gf",
            "girl propose boy",
        ],
    }
    faq_page = {"@context": "https://schema.org", "@type": "FAQPage", "mainEntity": faq_schema}
    content += (
        "\n\n<!-- wp:html -->\n"
        f'<script type="application/ld+json">{json.dumps(faq_page, ensure_ascii=False)}</script>\n'
        f'<script type="application/ld+json">{json.dumps(article_schema, ensure_ascii=False)}</script>\n'
        "<!-- /wp:html -->\n"
    )

    post = api(
        "POST",
        "posts",
        {
            "title": TITLE,
            "slug": SLUG,
            "status": "publish",
            "author": 270271338,
            "featured_media": hero_media,
            "content": content,
            "excerpt": META_DESC,
            "meta": {
                "_yoast_wpseo_focuskw": FOCUS_KW,
                "_yoast_wpseo_title": YOAST_TITLE,
                "_yoast_wpseo_metadesc": META_DESC,
            },
        },
    )
    print("published", post["id"], post["link"])

    Path(ROOT / "output/Week1_Rank3_Propose_product_media.json").write_text(
        json.dumps(product_media, indent=2)
    )
    update_xlsx(post["id"], post["link"])

    html = urllib.request.urlopen(
        urllib.request.Request(post["link"] + "?v=audit", headers={"User-Agent": "Mozilla/5.0"}),
        timeout=30,
    ).read().decode("utf-8", "replace")
    checks = {
        "H1": len(re.findall(r"<h1\b", html, re.I)) == 1,
        "FAQ": "Frequently Asked Questions" in html,
        "carousel": "bs-cf-propose" in html,
        "buy now": html.count("Buy now") >= 6,
        "Vikas": "Vikas" in html,
    }
    print("audit", checks)


if __name__ == "__main__":
    main()

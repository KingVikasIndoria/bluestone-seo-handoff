#!/usr/bin/env python3
"""
Batch queue + live progress for Week 1-2 festive blogs (New publish only).

This does NOT fully auto-generate articles by itself (drafting + Higgsfield MCP
need the Cursor agent). It:

1. Loads the next N ranks from SEO Strategy 2026.xlsx
2. Tracks status in output/batch_queue.json
3. Prints / watches progress in the terminal
4. Hands the agent the next pending rank via `next-prompt`

Usage:
  python3 scripts/batch_blog_queue.py init --start 25 --count 10
  python3 scripts/batch_blog_queue.py status
  python3 scripts/batch_blog_queue.py watch
  python3 scripts/batch_blog_queue.py next-prompt
  python3 scripts/batch_blog_queue.py mark --rank 25 --status done --url URL --wp-id 30280
  python3 scripts/batch_blog_queue.py estimate
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path("/Users/vikasindoria/Documents/Geo and Seo/SEO Strategy 2026.xlsx")
QUEUE_PATH = ROOT / "output" / "batch_queue.json"
LOG_PATH = ROOT / "output" / "batch_queue.log"

# Minutes per blog based on Rank 22–24 runs (text + carousel + Type 3 + patch)
MIN_PER_BLOG = 12
MAX_PER_BLOG = 18  # measured Rank25 ~10.5 min; buffer for gift/education + retries


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def log(msg: str) -> None:
    line = f"[{utc_now()}] {msg}"
    print(line, flush=True)
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with LOG_PATH.open("a", encoding="utf-8") as fh:
        fh.write(line + "\n")


def load_queue() -> dict:
    if not QUEUE_PATH.exists():
        raise SystemExit(f"No queue yet. Run: python3 scripts/batch_blog_queue.py init")
    return json.loads(QUEUE_PATH.read_text(encoding="utf-8"))


def save_queue(data: dict) -> None:
    data["updated_at"] = utc_now()
    QUEUE_PATH.parent.mkdir(parents=True, exist_ok=True)
    QUEUE_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def read_plan_rows(start: int, count: int) -> list[dict]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Week 1-2"]
    headers = {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }
    rows = []
    for row in range(2, ws.max_row + 1):
        rank = ws.cell(row, headers["Rank"]).value
        if not isinstance(rank, (int, float)):
            continue
        rank = int(rank)
        if rank < start:
            continue
        if len(rows) >= count:
            break
        old_url = ws.cell(row, headers["Bluestone Blog URL"]).value
        rows.append(
            {
                "rank": rank,
                "sheet_row": row,
                "action": "New",
                "primary_kw": ws.cell(row, headers["Primary Keyword"]).value,
                "supporting_kws": ws.cell(row, headers["Supporting Keywords"]).value,
                "volume": ws.cell(row, headers["Volume"]).value,
                "kd": ws.cell(row, headers["KD"]).value,
                "suggested_slug": ws.cell(row, headers["Suggested URL Slug"]).value,
                "competitor": ws.cell(row, headers["CaratLane URL"]).value,
                "old_bluestone_url_reference_only": old_url,
                "status": "pending",
                "live_url": None,
                "wp_post_id": None,
                "started_at": None,
                "finished_at": None,
                "type3_products": None,
                "note": None,
                "error": None,
            }
        )
    return rows


def cmd_init(start: int, count: int) -> None:
    items = read_plan_rows(start, count)
    if not items:
        raise SystemExit("No ranks found in that range.")
    data = {
        "created_at": utc_now(),
        "updated_at": utc_now(),
        "policy": "New blogs only. Ignore Optimize. Do not edit old BlueStone URLs.",
        "estimate_minutes_per_blog": {"min": MIN_PER_BLOG, "max": MAX_PER_BLOG},
        "items": items,
    }
    save_queue(data)
    log(f"Initialized queue: Ranks {items[0]['rank']}–{items[-1]['rank']} ({len(items)} blogs)")
    cmd_estimate()
    cmd_status()


def counts(data: dict) -> dict:
    c = {"pending": 0, "in_progress": 0, "done": 0, "failed": 0, "skipped": 0}
    for item in data["items"]:
        c[item["status"]] = c.get(item["status"], 0) + 1
    return c


def cmd_status() -> None:
    data = load_queue()
    c = counts(data)
    total = len(data["items"])
    done = c.get("done", 0)
    print()
    print("=" * 72)
    print(f" BATCH QUEUE  |  {done}/{total} done  |  updated {data.get('updated_at')}")
    print("=" * 72)
    for item in data["items"]:
        mark = {
            "pending": "·",
            "in_progress": "▶",
            "done": "✓",
            "failed": "✗",
            "skipped": "–",
        }.get(item["status"], "?")
        kw = (item.get("primary_kw") or "")[:42]
        url = item.get("live_url") or ""
        print(f"  [{mark}] Rank {item['rank']:<3}  {item['status']:<12}  {kw}")
        if url:
            print(f"           {url}")
        if item.get("error"):
            print(f"           ERROR: {item['error']}")
    print("-" * 72)
    print(
        f"  pending={c.get('pending',0)}  in_progress={c.get('in_progress',0)}  "
        f"done={c.get('done',0)}  failed={c.get('failed',0)}"
    )
    print("=" * 72)
    print()


def cmd_estimate() -> None:
    data = load_queue()
    pending = sum(1 for i in data["items"] if i["status"] in ("pending", "in_progress", "failed"))
    done = sum(1 for i in data["items"] if i["status"] == "done")
    remaining = sum(1 for i in data["items"] if i["status"] in ("pending", "failed"))
    print()
    print("TIME ESTIMATE (sequential, one blog at a time)")
    print(f"  Per blog:     ~{MIN_PER_BLOG}–{MAX_PER_BLOG} minutes")
    print(f"  Done so far:  {done}")
    print(f"  Remaining:    {remaining}")
    print(
        f"  Remaining ETA: ~{remaining * MIN_PER_BLOG // 60}h {remaining * MIN_PER_BLOG % 60}m"
        f"  to  ~{remaining * MAX_PER_BLOG // 60}h {remaining * MAX_PER_BLOG % 60}m"
    )
    print(
        f"  Full batch ({len(data['items'])}): "
        f"~{len(data['items']) * MIN_PER_BLOG // 60}h {len(data['items']) * MIN_PER_BLOG % 60}m"
        f"  to  ~{len(data['items']) * MAX_PER_BLOG // 60}h {len(data['items']) * MAX_PER_BLOG % 60}m"
    )
    print()
    print("Breakdown per blog:")
    print("  Competitor + brief + draft     ~12–18 min")
    print("  Products + carousel + WP text  ~8–12 min")
    print("  Type 3 Higgsfield + QA/patch   ~10–18 min (retries add time)")
    print("  Checklist + xlsx update        ~2–3 min")
    print()


def cmd_watch(interval: float = 5.0) -> None:
    print(f"Watching {QUEUE_PATH} (Ctrl+C to stop). Refresh every {interval:.0f}s.\n")
    last = None
    try:
        while True:
            data = load_queue()
            snap = json.dumps(
                [(i["rank"], i["status"], i.get("live_url")) for i in data["items"]],
                ensure_ascii=False,
            )
            if snap != last:
                # clear-ish view
                print("\033[H\033[J", end="")
                cmd_status()
                cmd_estimate()
                if LOG_PATH.exists():
                    lines = LOG_PATH.read_text(encoding="utf-8").splitlines()[-8:]
                    print("RECENT LOG")
                    for line in lines:
                        print(" ", line)
                    print()
                last = snap
            time.sleep(interval)
    except KeyboardInterrupt:
        print("\nStopped watching.")


def next_pending(data: dict) -> dict | None:
    for item in data["items"]:
        if item["status"] in ("pending", "failed"):
            return item
        if item["status"] == "in_progress":
            return item
    return None


def cmd_next_prompt() -> None:
    data = load_queue()
    item = next_pending(data)
    if not item:
        print("All ranks in this batch are done.")
        return
    if item["status"] != "in_progress":
        item["status"] = "in_progress"
        item["started_at"] = utc_now()
        save_queue(data)
        log(f"Rank {item['rank']} marked in_progress")

    old = item.get("old_bluestone_url_reference_only")
    old_line = (
        f"- Do NOT edit existing BlueStone URL (reference only): {old}\n"
        if old
        else "- No existing BlueStone URL on sheet.\n"
    )
    prompt = f"""Continue the festive batch queue. Process ONLY the next pending rank end-to-end, then stop and update the queue.

Read fully first:
- article generation seo codex/HANDOFF.md
- KnowledgeBase/Writing/SOP_ARTICLE_GENERATION.md
- KnowledgeBase/Writing/ARTICLE_WORKFLOW.md
- KnowledgeBase/Writing/HIGGSFIELD_IMAGE_GENERATION.md (Image SEO + people-first Type 3 + credits)
- KnowledgeBase/Writing/Blog-SEO-AEO-GEO-Checklist-v2.md
- .cursor/rules/new-blogs-only.mdc
- .cursor/rules/carousel-seo-images.mdc
- .cursor/rules/type3-fair-skinned-indians.mdc
- .cursor/rules/image-seo.mdc
- article generation seo codex/output/batch_queue.json

Next article: SEO Strategy 2026.xlsx → Week 1-2 → Rank {item['rank']}
- Primary KW: {item['primary_kw']}
- Action: NEW only (ignore Optimize)
{old_line}- Volume / KD: {item.get('volume')} / {item.get('kd')}
- Supporting KWs: {item.get('supporting_kws')}
- Competitor: {item.get('competitor')}
- Suggested slug: {item.get('suggested_slug')} (override with locked festive year)
- Category: Festive Wishes (+ Quotes & Wishes). Never Uncategorized.

Execute full New publish (adapt publish_gudi_padwa / Rank 23–24 patterns):
1. Competitor pass → output/Week1_Rank{item['rank']}_*_Competitor_Analysis.md
2. Brief + draft NEW pillar (100+ lines if title claims 100+)
3. Pick 6 products from Seo Products CSV
4. Type 2 carousel from ProductImages/seo images/ only → WebP + Image SEO
5. Publish NEW WP post (author Vikas, ID 270271338)
6. FAQs 40–80 words + FAQPage + BlogPosting
7. Type 3 AFTER text publish — exactly 3 images (hero+flatlay+lifestyle), count:1 each
8. Checklist v2; meta title ≤60; meta desc ~150–160; no old years in body
9. Update Week 1-2 Rank {item['rank']} in SEO Strategy 2026.xlsx
10. Update batch queue:
    python3 scripts/batch_blog_queue.py mark --rank {item['rank']} --status done --url LIVE_URL --wp-id ID --type3-products "Hero: CODE Name | Flatlay: CODE Name | Lifestyle: CODE Name"
11. Print live URL + Type 3 products used

Hard rules:
- No prices; products only from Seo Products CSV
- Carousel mid-article, never last
- No em/en dashes; no "word - word"
- Type 3: scene-first people prompts; 2 product refs max; fair-skinned Indians; jewellery worn/on surface; QA against real product refs before shipping
- Do not redo earlier completed ranks unless asked

Env: .env WP_USER / WP_APP_PASSWORD; Higgsfield via Cursor MCP
"""
    out = ROOT / "output" / "batch_next_agent_prompt.txt"
    out.write_text(prompt, encoding="utf-8")
    print(prompt)
    print(f"\n(Also saved to {out})")


def cmd_mark(rank: int, status: str, url: str | None, wp_id: int | None, type3: str | None, note: str | None, error: str | None) -> None:
    data = load_queue()
    item = next((i for i in data["items"] if i["rank"] == rank), None)
    if not item:
        raise SystemExit(f"Rank {rank} not in queue")
    item["status"] = status
    if url:
        item["live_url"] = url
    if wp_id:
        item["wp_post_id"] = wp_id
    if type3:
        item["type3_products"] = type3
    if note:
        item["note"] = note
    if error:
        item["error"] = error
    if status == "in_progress" and not item.get("started_at"):
        item["started_at"] = utc_now()
    if status in ("done", "failed", "skipped"):
        item["finished_at"] = utc_now()
    save_queue(data)
    log(f"Rank {rank} -> {status}" + (f" {url}" if url else ""))
    cmd_status()


def main() -> None:
    parser = argparse.ArgumentParser(description="Festive blog batch queue + progress")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_init = sub.add_parser("init", help="Create queue from xlsx")
    p_init.add_argument("--start", type=int, default=25)
    p_init.add_argument("--count", type=int, default=10)

    sub.add_parser("status", help="Show progress once")
    sub.add_parser("estimate", help="Show time estimate")
    p_watch = sub.add_parser("watch", help="Live progress in terminal")
    p_watch.add_argument("--interval", type=float, default=5.0)
    sub.add_parser("next-prompt", help="Print agent prompt for next pending rank")

    p_mark = sub.add_parser("mark", help="Update rank status")
    p_mark.add_argument("--rank", type=int, required=True)
    p_mark.add_argument("--status", required=True, choices=["pending", "in_progress", "done", "failed", "skipped"])
    p_mark.add_argument("--url", default=None)
    p_mark.add_argument("--wp-id", type=int, default=None)
    p_mark.add_argument("--type3-products", default=None)
    p_mark.add_argument("--note", default=None)
    p_mark.add_argument("--error", default=None)

    args = parser.parse_args()
    if args.cmd == "init":
        cmd_init(args.start, args.count)
    elif args.cmd == "status":
        cmd_status()
    elif args.cmd == "estimate":
        cmd_estimate()
    elif args.cmd == "watch":
        cmd_watch(args.interval)
    elif args.cmd == "next-prompt":
        cmd_next_prompt()
    elif args.cmd == "mark":
        cmd_mark(args.rank, args.status, args.url, args.wp_id, args.type3_products, args.note, args.error)


if __name__ == "__main__":
    main()

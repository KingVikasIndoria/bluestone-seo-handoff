#!/usr/bin/env python3
"""Audit fixes for Rank 21 Ugadi NEW post WP#30229.

Keeps existing Telugu lines from live HTML (UTF-8), expands with English extras,
removes old years, fixes FAQs/meta/category/image SEO/Gutenberg headings.
"""
from __future__ import annotations

import base64
import json
import os
import re
import urllib.request
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
POST_ID = 30229
CAT_FESTIVE = 554493477
CAT_QUOTES = 554493415
API = "https://blog.bluestone.com/wp-json/wp/v2"


def load_env() -> None:
    for line in (ROOT / ".env").read_text(encoding="utf-8").splitlines():
        if "=" in line and not line.strip().startswith("#"):
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip("'\""))


def api(method: str, path: str, payload: dict | None = None):
    data = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        f"{API}{path}",
        data=data,
        method=method,
        headers={
            "Authorization": "Basic "
            + base64.b64encode(
                f"{os.environ['WP_USER']}:{os.environ['WP_APP_PASSWORD']}".encode()
            ).decode(),
            "User-Agent": "ugadi-audit-fix",
            "Content-Type": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=120) as resp:
        return json.loads(resp.read().decode("utf-8"))


def extract_list_items(block: str) -> list[str]:
    return re.findall(r"<li>([\s\S]*?)</li>", block)


def ol(items: list[str]) -> str:
    lis = "\n".join(f"<li>{x}</li>" for x in items)
    return (
        '<!-- wp:list {"ordered":true} -->\n'
        f'<ol class="wp-block-list">\n{lis}\n</ol>\n<!-- /wp:list -->'
    )


def ul(items: list[str]) -> str:
    lis = "\n".join(f"<li>{x}</li>" for x in items)
    return f'<!-- wp:list -->\n<ul class="wp-block-list">\n{lis}\n</ul>\n<!-- /wp:list -->'


def h2(t: str) -> str:
    return f'<!-- wp:heading -->\n<h2 class="wp-block-heading">{t}</h2>\n<!-- /wp:heading -->'


def h3(t: str) -> str:
    return (
        '<!-- wp:heading {"level":3} -->\n'
        f'<h3 class="wp-block-heading">{t}</h3>\n<!-- /wp:heading -->'
    )


def p(t: str) -> str:
    return f"<!-- wp:paragraph -->\n<p>{t}</p>\n<!-- /wp:paragraph -->"


def img(mid: int, src: str, alt: str) -> str:
    return (
        f'<!-- wp:image {{"id":{mid},"sizeSlug":"full","linkDestination":"none"}} -->\n'
        f'<figure class="wp-block-image size-full">'
        f'<img src="{src}" alt="{alt}" class="wp-image-{mid}"/></figure>\n'
        f"<!-- /wp:image -->"
    )


def main() -> None:
    load_env()
    live = api("GET", f"/posts/{POST_ID}?context=edit")
    raw = live["content"]["raw"]

    car_m = re.search(
        r"<!-- wp:html -->\s*<style>\s*\n\.bs-cf[\s\S]*?</script>\s*<!-- /wp:html -->",
        raw,
    )
    if not car_m:
        raise SystemExit("carousel block not found")
    carousel = car_m.group(0)
    for old, new in [
        (
            'alt="Happy Ugadi wishes gift idea: The Rohal Huggie Earrings from BlueStone"',
            'alt="Happy Ugadi wishes in Telugu 2027 gift idea: The Rohal Huggie Earrings"',
        ),
        (
            'alt="Happy Ugadi wishes gift idea: The Valeria Rose Pendant from BlueStone"',
            'alt="Happy Ugadi wishes in Telugu 2027 gift idea: The Valeria Rose Pendant"',
        ),
        (
            'alt="Happy Ugadi wishes gift idea: The Muricelle Bangle from BlueStone"',
            'alt="Happy Ugadi wishes in Telugu 2027 gift idea: The Muricelle Bangle"',
        ),
        (
            'alt="Happy Ugadi wishes gift idea: The Gigi Ring from BlueStone"',
            'alt="Happy Ugadi wishes in Telugu 2027 gift idea: The Gigi Ring"',
        ),
        (
            'alt="Happy Ugadi wishes gift idea: The Aarabhi Mangalsutra from BlueStone"',
            'alt="Happy Ugadi wishes in Telugu 2027 gift idea: The Aarabhi Mangalsutra"',
        ),
        (
            'alt="Happy Ugadi wishes gift idea: The Asya Huggie Earrings from BlueStone"',
            'alt="Happy Ugadi wishes in Telugu 2027 gift idea: The Asya Huggie Earrings"',
        ),
    ]:
        carousel = carousel.replace(old, new)

    ols = re.findall(r'<ol class="wp-block-list">([\s\S]*?)</ol>', raw)
    uls = re.findall(r'<ul class="wp-block-list">([\s\S]*?)</ul>', raw)
    if len(ols) < 6 or not uls:
        raise SystemExit(f"unexpected list counts ols={len(ols)} uls={len(uls)}")

    # Idempotent: if already expanded (>=100 lis), skip English appends; still rebuild structure.
    existing_li = len(re.findall(r"<li>", raw))
    expand = existing_li < 100

    def maybe_extra(base: list[str], extras: list[str]) -> list[str]:
        if not expand:
            # dedupe only
            seen: set[str] = set()
            out: list[str] = []
            for it in base:
                key = re.sub(r"\s+", " ", it).strip().lower()
                if key in seen:
                    continue
                seen.add(key)
                out.append(it)
            return out
        return base + extras

    whatsapp = maybe_extra(
        extract_list_items(ols[0]),
        [
            "Happy Ugadi 2027: wishing your home peace, health, and a bright Telugu New Year morning.",
            "Quick WhatsApp line: Happy Ugadi 2027. Fresh hopes, warm hearts, sweet beginnings.",
            "Sending Ugadi 2027 wishes across the group chat with love and gratitude.",
            "May this Ugadi bring calm mornings, kind words, and goals that feel closer.",
            "Short English WhatsApp wish: Happy Ugadi 2027. Celebrate with family and hope.",
            "Ugadi 2027 greeting: May mango leaves and warm smiles welcome your new year.",
            "Copy-ready: Happy Ugadi wishes for 2027. Stay healthy, stay joyful, stay close.",
            "One-line status: Happy Ugadi 2027 | New year energy for everyone at home.",
        ],
    )
    family = maybe_extra(
        extract_list_items(ols[1]),
        [
            "Dear family, Happy Ugadi 2027. May our home stay full of prayer, food, and laughter.",
            "To every elder at home: Happy Ugadi 2027. Your blessings shape our year.",
            "For siblings this Ugadi: thank you for the jokes, support, and shared memories.",
            "Family card line: Happy Ugadi 2027. Together is our favourite festival tradition.",
            "Parents and kids together: Happy Ugadi 2027. May this year feel lighter and kinder.",
            "Whole-house wish: Happy Ugadi 2027. Peace in the morning, joy through the day.",
            "Ugadi 2027 for our family WhatsApp group: more meetings, more smiles, more gratitude.",
        ],
    )
    friends = maybe_extra(
        extract_list_items(ols[2]),
        [
            "Happy Ugadi 2027, friend. May this Telugu New Year bring good news your way.",
            "To my closest people: Happy Ugadi. Grateful for every shared celebration.",
            "College friends reunion wish: Happy Ugadi 2027. Same laughs, newer dreams.",
            "Office buddy note: Happy Ugadi! Wishing you a calmer, brighter year ahead.",
            "Long-distance friend: Happy Ugadi 2027. Miles apart, still celebrating with you.",
            "Group chat line: Happy Ugadi everyone. New year, same wonderful circle.",
        ],
    )
    english = maybe_extra(
        extract_list_items(ols[3]),
        [
            "Happy Ugadi 2027 to you and your family. May every prayer find a gentle answer.",
            "Wishing you a balanced year ahead, just like Ugadi Pachadi: sweet, bold, and full of life.",
            "May Ugadi 2027 open doors to better health, kinder days, and quieter joy.",
            "English office note: Happy Ugadi! Wishing your team a fresh and successful Telugu New Year.",
            "For cards: Happy Ugadi 2027. New leaves, new light, new love at home.",
            "Simple and warm: Happy Ugadi wishes for 2027. Celebrate beginnings with gratitude.",
        ],
    )
    traditional = maybe_extra(
        extract_list_items(ols[4]),
        [
            "May the first sunrise of Ugadi 2027 fill your home with devotion and hope.",
            "Traditional English blessing: Welcome Ugadi with prayer, gratitude, and a calm heart.",
            "Let Ugadi Pachadi remind you to meet every season of life with grace.",
            "Spiritual wish for 2027: May faith, family, and kindness lead your new year.",
            "May this Ugadi open with light in the home and peace in every prayer.",
            "A calm New Year blessing: Happy Ugadi 2027. Stay grounded, stay grateful.",
        ],
    )
    funny = maybe_extra(
        extract_list_items(ols[5]),
        [
            "Happy Ugadi 2027! May sweets arrive before meetings. Priorities matter.",
            "May your group chat stay funnier than your festival outfit stress.",
            "Happy Ugadi! May relatives praise the food before they ask about your plans.",
            "New year wish: less traffic, more laddoos. Happy Ugadi 2027!",
            "Festival mode on: Happy Ugadi. Selfie first, resolutions later.",
        ],
    )
    captions = maybe_extra(
        extract_list_items(uls[0]),
        [
            "Simple status: Happy Ugadi 2027. New year, soft heart.",
            "Family frame: Ugadi wishes | Our home, our hope, our year.",
            "Festive reel line: Toran up, Pachadi ready, hearts full.",
            "Caption for friends: Happy Ugadi wishes in Telugu and English, same love.",
            "Story sticker idea: Happy Ugadi 2027 | Fresh year, grateful heart.",
        ],
    )

    faqs = [
        (
            "What are some happy Ugadi wishes in Telugu for 2027?",
            "Start with a clear greeting such as a short Telugu line that wishes joy, health, and success for the family. Happy Ugadi wishes in Telugu work best when they mention family, health, and a fresh year. Add a personal name or relation if you can, then close with one hope for Ugadi 2027. Short, warm lines are easiest to share on WhatsApp and festive cards.",
        ),
        (
            "What is a good WhatsApp Ugadi wish in Telugu?",
            "Keep WhatsApp Ugadi wishes under three lines: a greeting, one blessing, and a warm close. A short Telugu line with Happy Ugadi 2027 shares cleanly in family groups, especially early on Ugadi morning. If the chat mixes languages, pair one Telugu line with a short English wish. Avoid long paragraphs that get lost in busy festival chats.",
        ),
        (
            "How do I write ugadi wishes in Telugu for family?",
            "Name the relation, thank them, and add one hope for the new year. Ugadi wishes in Telugu for 2027 feel warmer when they sound personal rather than copied from a random forward. For elders, keep the tone respectful. For kids, keep it light and joyful. One sincere sentence often means more than a long message.",
        ),
        (
            "What are happy Ugadi wishes 2027 in English?",
            "Try: Happy Ugadi 2027! May this Telugu New Year fill your home with peace, health, and fresh beginnings. English Ugadi 2027 wishes help when your chat has mixed-language friends or colleagues. You can also mention Ugadi Pachadi, mango leaves, or a simple prayer for balance. Keep the message short enough for a card or status update.",
        ),
        (
            "When should I send Ugadi wishes in 2027?",
            "Send happy Ugadi wishes in Telugu early on Ugadi morning, ideally after a short prayer or family greeting. A simple order that feels respectful is elders first, then family groups, then friends. If someone is travelling, a short message before sunrise still works. For social captions, post once your home toran or festive table is ready.",
        ),
        (
            "What is a thoughtful Ugadi gift idea from BlueStone?",
            "Earrings, pendants, bangles, rings, and mangalsutra styles make lasting Ugadi keepsakes that go beyond sweets and cards. Designs such as The Rohal Huggie Earrings, The Valeria Rose Pendant, or The Aarabhi Mangalsutra suit festive new beginnings. Choose a piece the person can wear after the celebration too. Keep the note short and warm with your Ugadi wish.",
        ),
        (
            "What does Ugadi Pachadi teach us for wishes?",
            "Ugadi Pachadi mixes sweet, sour, and bitter tastes to mirror real life. A thoughtful Ugadi message can wish someone balance through every flavour of the year ahead. That is why many Telugu greetings mention Pachadi, mango leaves, and a calm new beginning. Use that idea when you want a traditional or spiritual tone for Ugadi 2027.",
        ),
    ]

    # Prefer original Telugu sample lines in FAQ #1 and #2 from live first items
    if whatsapp:
        faqs[0] = (
            faqs[0][0],
            f"Start with a clear greeting such as “{whatsapp[0]}” Happy Ugadi wishes in Telugu work best when they mention family, health, and a fresh year. Add a personal name or relation if you can, then close with one hope for Ugadi 2027. Short, warm lines are easiest to share on WhatsApp and festive cards.",
        )
    if len(whatsapp) > 1:
        faqs[1] = (
            faqs[1][0],
            f"Keep WhatsApp Ugadi wishes under three lines: a greeting, one blessing, and a warm close. Example: “{whatsapp[1]}” Short Telugu lines share cleanly in family groups, especially early on Ugadi morning. If the chat mixes languages, pair one Telugu line with a short English wish. Avoid long paragraphs that get lost in busy festival chats.",
        )
    if family:
        faqs[2] = (
            faqs[2][0],
            f"Name the relation, thank them, and add one hope for the new year. Example for parents: “{family[0]}” Ugadi wishes in Telugu for 2027 feel warmer when they sound personal. For elders, keep the tone respectful. For kids, keep it light and joyful.",
        )

    for q, a in faqs:
        wc = len(a.split())
        assert 40 <= wc <= 85, (q, wc)

    total = (
        len(whatsapp)
        + len(family)
        + len(friends)
        + len(english)
        + len(traditional)
        + len(funny)
        + len(captions)
    )
    print("total wish/caption lines", total)
    assert total >= 100, total

    title = "100+ Happy Ugadi Wishes in Telugu for 2027 | WhatsApp Messages"
    seo_title = "Happy Ugadi Wishes in Telugu 2027 | BlueStone"
    meta_desc = (
        "Copy 100+ happy Ugadi wishes in Telugu for 2027, plus WhatsApp lines, "
        "family messages, English greetings, captions, and FAQs ready to share today."
    )
    assert len(seo_title) <= 60, len(seo_title)
    assert 140 <= len(meta_desc) <= 165, (len(meta_desc), meta_desc)

    alt_flat = (
        "Happy Ugadi wishes in Telugu 2027 flatlay with The Muricelle Bangle and mango leaves"
    )
    alt_life = (
        "Happy Ugadi wishes in Telugu 2027 lifestyle, The Rohal Huggie Earrings "
        "on fair-skinned Indian woman"
    )
    alt_hero = (
        "Happy Ugadi wishes in Telugu 2027 hero, fair-skinned Indian woman "
        "wearing The Valeria Rose Pendant"
    )

    parts = [
        "<!-- wp:html -->",
        "<style>",
        ".bs-eeat{margin:0 auto 1.25rem;max-width:720px;text-align:center;font-size:.95rem;color:#444;line-height:1.5}",
        ".bs-eeat strong{color:#111}",
        ".entry-content img,.wp-block-image img{max-width:100%;height:auto}",
        "</style>",
        "<!-- /wp:html -->",
        "",
        '<!-- wp:paragraph {"align":"center"} -->',
        '<p class="has-text-align-center bs-eeat">By <strong>Vikas</strong>, BlueStone Editorial</p>',
        "<!-- /wp:paragraph -->",
        "",
        p(
            "Looking for <strong>happy Ugadi wishes in Telugu</strong> for WhatsApp, "
            "family chats, and festive cards? This guide gathers Telugu Ugadi greetings, "
            "English Ugadi 2027 messages, captions, and traditional blessings so you can "
            "copy a line that fits every loved one."
        ),
        "",
        p(
            "<strong>TL;DR:</strong> Send a short Telugu WhatsApp wish, a warm family message, "
            "an English Ugadi 2027 greeting, or a caption for status updates. Every list is "
            "refreshed for Ugadi 2027."
        ),
        "",
        p(
            "Ugadi marks the Telugu New Year with mango leaves, Ugadi Pachadi, and prayers "
            "for a balanced year. A sincere wish, shared early in the morning, can turn the "
            "festival into a memory your family keeps all year."
        ),
        "",
        h2("Short WhatsApp Ugadi Wishes in Telugu"),
        "",
        p(
            "These copy-ready WhatsApp Ugadi wishes keep your greeting short, warm, and "
            "easy to paste into family groups."
        ),
        "",
        ol(whatsapp),
        "",
        h2("Happy Ugadi 2027 Wishes in Telugu for Family"),
        "",
        p(
            "Share these heartfelt happy Ugadi wishes in Telugu with parents, siblings, "
            "elders, and children at home."
        ),
        "",
        ol(family),
        "",
        h2("Ugadi Wishes in Telugu for Friends"),
        "",
        p(
            "Send these ugadi wishes in Telugu to friends who feel like family, whether "
            "they are nearby or far away."
        ),
        "",
        ol(friends),
        "",
        img(
            30236,
            "https://blog.bluestone.com/wp-content/uploads/2026/07/ugadi-flatlay-2027-1.webp",
            alt_flat,
        ),
        "",
        h2("Happy Ugadi Wishes 2027 in English"),
        "",
        p(
            "Need English lines for mixed chats? Use these happy Ugadi wishes 2027 messages "
            "for cards, office groups, and captions."
        ),
        "",
        ol(english),
        "",
        h2("A Soft Ugadi Gift Idea (If You Are Gifting Too)"),
        "",
        p(
            "Ugadi is a beautiful time to mark new beginnings with something lasting. "
            "Earrings, pendants, bangles, rings, and mangalsutra styles make thoughtful "
            "festive keepsakes. Explore these six approved designs from the BlueStone collection."
        ),
        "",
        carousel,
        "",
        img(
            30237,
            "https://blog.bluestone.com/wp-content/uploads/2026/07/ugadi-lifestyle-2027-1.webp",
            alt_life,
        ),
        "",
        h2("Ugadi Captions &amp; Status for Instagram and WhatsApp"),
        "",
        p(
            "Pair these Ugadi captions with a home photo, mango toran reel, or simple "
            "family portrait."
        ),
        "",
        ul(captions),
        "",
        h2("Traditional and Spiritual Ugadi Blessings"),
        "",
        p(
            "Use these traditional lines when you want Ugadi Pachadi wisdom, prayerful tone, "
            "or a calmer New Year blessing."
        ),
        "",
        ol(traditional),
        "",
        h2("Funny Ugadi Wishes"),
        "",
        p("Lighten the group chat with these playful Ugadi lines for friends and siblings."),
        "",
        ol(funny),
        "",
        h2("How to Pick the Right Ugadi Message"),
        "",
        p(
            "Choose Telugu for grandparents and family groups, English for mixed-language "
            "friends, and a short WhatsApp line when you are greeting many people at once. "
            "Personalize with a name, keep the tone warm, and use Ugadi 2027 in any "
            "year-sensitive greeting before you send it."
        ),
        "",
        h2("More Festive &amp; Occasion Reads"),
        "",
        p(
            'Explore our other guides including <a href="https://blog.bluestone.com/happy-holi-wishes-messages-quotes-2027/">Holi wishes for 2027</a>, '
            '<a href="https://blog.bluestone.com/happy-diwali-wishes-messages-quotes-2026/">Diwali wishes for 2026</a>, '
            '<a href="https://blog.bluestone.com/ganesh-chaturthi-wishes-in-english/">Ganesh Chaturthi wishes for 2026</a>, '
            '<a href="https://blog.bluestone.com/makar-sankranti-quotes-wishes-pongal-and-magh-bihu-festival-greetings/">Pongal wishes in Tamil for 2027</a>, '
            '<a href="https://blog.bluestone.com/dussehra-wishes-in-english/">Dussehra wishes for 2026</a>, and '
            '<a href="https://blog.bluestone.com/bhai-dooj-wishes/">Bhai Dooj wishes for 2026</a>. '
            'Learn more about <a href="https://en.wikipedia.org/wiki/Ugadi">Ugadi on Wikipedia</a>.'
        ),
        "",
        h2("Frequently Asked Questions about Happy Ugadi Wishes in Telugu"),
        "",
    ]
    for q, a in faqs:
        parts += [h3(q), "", p(a), ""]
    parts += [
        h2("Conclusion"),
        "",
        p(
            "The best happy Ugadi wishes in Telugu sound personal, hopeful, and ready to share. "
            "Pick a WhatsApp line, bless your family, and start Ugadi 2027 with warmth. "
            "Happy Ugadi 2027!"
        ),
        "",
    ]

    faq_schema = {
        "@context": "https://schema.org",
        "@type": "FAQPage",
        "mainEntity": [
            {
                "@type": "Question",
                "name": q,
                "acceptedAnswer": {"@type": "Answer", "text": a},
            }
            for q, a in faqs
        ],
    }
    blog_schema = {
        "@context": "https://schema.org",
        "@type": "BlogPosting",
        "headline": title,
        "description": meta_desc,
        "datePublished": "2026-07-19",
        "dateModified": date.today().isoformat(),
        "author": {"@type": "Person", "name": "Vikas"},
        "publisher": {
            "@type": "Organization",
            "name": "BlueStone",
            "url": "https://www.bluestone.com/",
        },
        "image": [
            "https://blog.bluestone.com/wp-content/uploads/2026/07/ugadi-hero-2027-1.webp",
            "https://blog.bluestone.com/wp-content/uploads/2026/07/ugadi-flatlay-2027-1.webp",
            "https://blog.bluestone.com/wp-content/uploads/2026/07/ugadi-lifestyle-2027-1.webp",
        ],
        "mainEntityOfPage": {
            "@type": "WebPage",
            "@id": "https://blog.bluestone.com/happy-ugadi-wishes-in-telugu-2027/",
        },
        "keywords": [
            "happy ugadi wishes in telugu",
            "happy ugadi 2027 wishes in telugu",
            "ugadi wishes in telugu",
            "whatsapp ugadi wishes",
            "happy ugadi wishes 2027",
            "ugadi 2027 wishes",
            "ugadi captions",
        ],
    }
    parts += [
        "<!-- wp:html -->",
        f'<script type="application/ld+json" id="bs-faq-schema">{json.dumps(faq_schema, ensure_ascii=False)}</script>',
        f'<script type="application/ld+json" id="bs-blogposting-schema">{json.dumps(blog_schema, ensure_ascii=False)}</script>',
        "<!-- /wp:html -->",
    ]

    content = "\n".join(parts)
    for y in ["2020", "2021", "2022", "2023", "2024", "2025"]:
        if y in content:
            raise SystemExit(f"old year {y} still present")
    if "<!-- wp:heading -->\n\n\n<!-- wp:image" in content:
        raise SystemExit("broken heading/image order still present")

    for mid, fields in {
        30235: {
            "alt_text": alt_hero,
            "title": "Happy Ugadi Wishes in Telugu 2027 Hero Image",
        },
        30236: {
            "alt_text": alt_flat,
            "title": "Happy Ugadi Wishes in Telugu 2027 Flatlay",
        },
        30237: {
            "alt_text": alt_life,
            "title": "Happy Ugadi Wishes in Telugu 2027 Lifestyle",
        },
    }.items():
        api("POST", f"/media/{mid}", fields)
        print("updated media", mid)

    out = api(
        "POST",
        f"/posts/{POST_ID}",
        {
            "title": title,
            "content": content,
            "categories": [CAT_FESTIVE, CAT_QUOTES],
            "meta": {
                "_yoast_wpseo_focuskw": "happy ugadi wishes in telugu",
                "_yoast_wpseo_title": seo_title,
                "_yoast_wpseo_metadesc": meta_desc,
            },
        },
    )
    print("updated post", out["id"], out["link"])
    print("categories", out.get("categories"))

    (ROOT / "output" / "Week1_Rank21_Ugadi_article.html").write_text(content, encoding="utf-8")

    p2 = api("GET", f"/posts/{POST_ID}?context=edit")
    c = p2["content"]["raw"]
    text = re.sub(r"<script[\s\S]*?</script>", "", c)
    text = re.sub(r"<[^>]+>", " ", text)
    print("old years", {y: text.count(y) for y in ["2021", "2022", "2023"]})
    print("li count", len(re.findall(r"<li>", c)))
    print("seo title", p2["meta"].get("_yoast_wpseo_title"), len(p2["meta"].get("_yoast_wpseo_title") or ""))
    print("meta len", len(p2["meta"].get("_yoast_wpseo_metadesc") or ""))
    print("cats", p2.get("categories"))
    print("broken heading?", "<!-- wp:heading -->\n\n\n<!-- wp:image" in c)

    from openpyxl import load_workbook

    xlsx = Path("/Users/vikasindoria/Documents/Geo and Seo/SEO Strategy 2026.xlsx")
    wb = load_workbook(xlsx)
    ws = wb["Week 1-2"]
    hh = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, hh["Rank"]).value == 21:
            note = ws.cell(r, hh["Execution Note"]).value or ""
            add = (
                " | Audit fixes 2026-07-19: Festive Wishes+Quotes cats, expanded 100+ lists, "
                "removed old years, FAQs 40-80w, meta title<=60, image alts/titles, Gutenberg heading fix"
            )
            if "Audit fixes 2026-07-19" not in note:
                ws.cell(r, hh["Execution Note"]).value = (note.rstrip() + add).strip()
            break
    wb.save(xlsx)
    print("xlsx updated")


if __name__ == "__main__":
    main()

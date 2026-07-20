#!/usr/bin/env python3
"""Publish Week 1 Rank 4: Happy Diwali Wishes, Messages & Quotes 2026."""
import base64
import json
import os
import re
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from html import escape

ROOT = Path(__file__).resolve().parents[1]
USER = os.environ["WP_USER"]
PWD = os.environ["WP_APP_PASSWORD"]
TOKEN = base64.b64encode(f"{USER}:{PWD}".encode()).decode()
AUTH = {"Authorization": f"Basic {TOKEN}", "User-Agent": "BluestoneSEO/1.0"}
API = "https://blog.bluestone.com/wp-json/wp/v2"

TITLE = "75+ Happy Diwali Wishes, Messages & Quotes for 2026"
SLUG = "happy-diwali-wishes-messages-quotes-2026"
META_DESC = (
    "Copy-ready Happy Diwali wishes, messages and quotes for 2026. Short WhatsApp lines, "
    "traditional greetings, family notes and captions to share light and joy this festival season."
)
FOCUS_KW = "adv happy diwali"
YOAST_TITLE = "Happy Diwali Wishes & Messages 2026 | BlueStone"

PRODUCTS = [
    {
        "name": "The Thyvarne Pendant",
        "url": "https://www.bluestone.com/pendants/the-thyvarne-pendant~173761.html",
        "png": ROOT / "ProductImages/seo images/Pendants/The Thyvarne Pendant.png",
    },
    {
        "name": "The Valeria Rose Pendant",
        "url": "https://www.bluestone.com/pendants/the-valeria-rose-pendant~181266.html",
        "png": ROOT / "ProductImages/seo images/Pendants/The Valeria Rose Pendant.png",
    },
    {
        "name": "The Asya Huggie Earrings",
        "url": "https://www.bluestone.com/earrings/the-asya-huggie-earrings~13494.html",
        "png": ROOT / "ProductImages/seo images/Earrings/The Asya Huggie Earrings.png",
    },
    {
        "name": "The Channing Bangle",
        "url": "https://www.bluestone.com/bangles/the-channing-bangle~975.html",
        "png": ROOT / "ProductImages/seo images/Bangles/The Channing Bangle.png",
    },
    {
        "name": "The Shining Star Bracelet",
        "url": "https://www.bluestone.com/bracelets/the-shining-star-bracelet~63731.html",
        "png": ROOT / "ProductImages/seo images/Bracelet/The Shining Star Bracelet.png",
    },
    {
        "name": "The Chevalier Gold Chain",
        "url": "https://www.bluestone.com/chains/the-chevalier-gold-chain~124914.html",
        "png": ROOT / "ProductImages/seo images/Chains/The Chevalier Gold Chain.png",
    },
]

TYPE3 = [
    {
        "key": "hero",
        "src": ROOT / "output/magnific_generated/diwali-hero-mood-2026.webp",
        "filename": "diwali-hero-mood-2026.webp",
        "alt": "Happy Diwali 2026 mood, diyas and marigold garlands with warm golden light",
        "featured": True,
    },
    {
        "key": "flatlay",
        "src": ROOT / "output/magnific_generated/diwali-flatlay-phone-2026.webp",
        "filename": "diwali-flatlay-phone-2026.webp",
        "alt": "Happy Diwali 2026 flatlay with Thyvarne pendant, blank phone, diyas and marigolds",
        "featured": False,
    },
    {
        "key": "lifestyle",
        "src": ROOT / "output/magnific_generated/diwali-lifestyle-family-2026.webp",
        "filename": "diwali-lifestyle-family-2026.webp",
        "alt": "Family celebrating Diwali with diyas, Happy Diwali wishes inspiration for 2026",
        "featured": False,
    },
]

SECTIONS = {
    "whatsapp": [
        "Happy Diwali 2026! May your home glow with diyas and your heart with peace.",
        "Wishing you light, laughter, and laddoos in equal measure. Shubh Diwali!",
        "May this festival bring prosperity to your doorstep and joy to your family.",
        "Sending warm Diwali wishes your way. Stay blessed and stay bright.",
        "Hope your Diwali sparkles as much as your smile. Happy festival of lights!",
        "May the diyas guide you toward success and calm. Shubh Deepavali 2026.",
        "Diwali Mubarak! May good fortune visit you and stay awhile.",
        "Light up the night, share the sweets, celebrate the people you love.",
        "Wishing you a Diwali filled with warmth, health, and happy reunions.",
        "May every diya you light remind you how loved you are. Happy Diwali!",
    ],
    "traditional": [
        "May Goddess Lakshmi bless your home with wealth, health, and harmony this Diwali.",
        "On this auspicious festival of lights, may victory of good over evil shine in your life.",
        "Wishing you and your family a joyous and prosperous Diwali celebration.",
        "May the divine light of Diwali dispel darkness and bring peace to your heart.",
        "Shubh Deepavali! May your year ahead be as bright as the festival sky.",
        "May this Diwali mark the beginning of new success and lasting happiness.",
        "Sending traditional Diwali blessings for prosperity and togetherness in 2026.",
        "May the warmth of diyas fill your home with eternal happiness and love.",
        "On Diwali, may your prayers rise like lanterns and your blessings multiply.",
        "Wishing you a sparkling Diwali and a year full of achievements ahead.",
    ],
    "heartfelt": [
        "Diwali feels brighter when I think of the people who make my life warm. That includes you.",
        "This festival reminds me that light wins when we choose kindness. Grateful for you.",
        "May your Diwali be wrapped in love, tied with laughter, and sealed with sweet memories.",
        "Some lights are not on balconies. They are in the hearts we share. Happy Diwali.",
        "I hope this Diwali gives you pause to feel proud of how far you have come.",
        "Wishing you quiet joy amid the noise, and real rest amid the celebrations.",
        "May the people you miss feel close tonight, even if miles apart.",
        "Diwali is a gentle nudge to tell people they matter. You matter to me.",
        "May your home feel full, your table feel generous, and your spirit feel light.",
        "Here is to old traditions, new beginnings, and the people who stay through both.",
    ],
    "inspirational": [
        "Let this Diwali inspire you to be the light someone else needs this year.",
        "Just as diyas push back the dark, may you push back doubt and keep going.",
        "The festival teaches us that small flames together create something magnificent.",
        "May Diwali 2026 be the chapter where hope feels practical, not distant.",
        "Light a lamp of gratitude tonight. It costs nothing and changes everything.",
        "Celebrate progress, not perfection. That is the spirit of renewal Diwali offers.",
        "May you release what weighed you down and make room for brighter choices.",
        "Diwali reminds us: even one steady light can guide a whole path forward.",
        "Share sweetness generously. Prosperity grows where generosity lives.",
        "This season, choose words that heal and actions that illuminate.",
    ],
    "lighthearted": [
        "May your Diwali be sweet, your selfies be bright, and your diet start tomorrow!",
        "Wishing you more diyas and fewer electricity bill surprises this year.",
        "Hope your rangoli stays intact and your siblings stay civil. Happy Diwali!",
        "Diwali rule: calories from mithai do not count. Science probably agrees.",
        "May your outfit slay and your firecracker budget stay intact.",
        "Sending wishes for a festival full of laughter, laddoos, and late-night chats.",
        "May your neighbours enjoy the lights and forgive the music volume.",
        "Diwali level unlocked: sweets eaten, family hugged, nap pending.",
        "Hope your Instagram feed glows harder than your phone screen tonight.",
        "Wishing you sparklers, spark, and zero rangoli disasters.",
    ],
    "short": [
        "Happy Diwali! Stay blessed.",
        "Shubh Deepavali 2026!",
        "Wishing you light and love.",
        "May your Diwali shine bright.",
        "Prosperity, peace, and joy to you.",
        "Happy festival of lights!",
        "Diwali Mubarak to you and yours.",
        "May good fortune find you tonight.",
        "Warm Diwali greetings your way.",
        "Celebrate bright, stay grateful.",
    ],
    "religious": [
        "May Lord Ganesha remove obstacles from your path this Diwali and always.",
        "Praying that Goddess Lakshmi fills your home with abundance and grace.",
        "May the blessings of Lord Rama guide you toward righteousness and peace.",
        "On this holy night, may divine light illuminate your mind and spirit.",
        "Wishing you spiritual growth and inner calm as you celebrate Diwali 2026.",
        "May your prayers on Diwali be heard and your heart be at rest.",
        "May the grace of the Almighty bring harmony to your home this festival.",
        "Diwali blessings for health, dharma, and devotion in the year ahead.",
        "May the sacred flame remind you that faith outlasts every season of doubt.",
        "Praying for peace in your home and compassion in every conversation.",
    ],
    "family": [
        "To my family: may our Diwali table be full and our hearts fuller still.",
        "Happy Diwali to the people who taught me what home feels like.",
        "May we gather, laugh loud, and create memories worth retelling next year.",
        "Wishing my parents health and ease as they light the first diya tonight.",
        "To my siblings: same chaos, same love, better sweets. Shubh Diwali!",
        "May our family bond grow stronger with every festival we share.",
        "Diwali wishes to the cousins, the aunties, and everyone who feels like family.",
        "Grateful for the traditions we keep and the new ones we are building together.",
        "May every member of our family find light, luck, and laughter this season.",
        "Here is to family photos, forced smiles, and genuinely happy endings.",
    ],
    "captions": [
        "Diya mode: on. Gratitude mode: also on.",
        "Gold lights, good people, great mithai.",
        "Diwali 2026 in one word: glowing.",
        "Serving festive looks and festive love.",
        "May your feed be as bright as your balcony.",
        "Shubh Deepavali from my home to yours.",
        "Little flames, big feelings.",
        "Celebrating light, legacy, and laddoos.",
        "Diwali nights hit different when family is near.",
        "Prosperity looks good on you. Happy Diwali!",
    ],
    "unique": [
        "May this Diwali open doors you stopped knocking on.",
        "Wishing you a festival where joy feels simple and sincere.",
        "May the glow of diyas mirror the glow of your next big win.",
        "Diwali reminder: share light, not just likes.",
        "May your year ahead sparkle with purpose, not just fireworks.",
    ],
}


def api(method, path, data=None, raw_body=None, headers=None):
    h = dict(AUTH)
    if headers:
        h.update(headers)
    if data is not None:
        h["Content-Type"] = "application/json"
        body = json.dumps(data).encode()
    else:
        body = raw_body
    req = urllib.request.Request(f"{API}/{path}", data=body, headers=h, method=method)
    with urllib.request.urlopen(req, timeout=180) as r:
        return json.loads(r.read().decode())


def to_webp(src: Path, dest: Path, size=None, carousel=False):
    from PIL import Image

    im = Image.open(src).convert("RGB")
    if carousel:
        tw, th = 960, 535
        im.thumbnail((tw, th), Image.Resampling.LANCZOS)
        canvas = Image.new("RGB", (tw, th), (245, 243, 240))
        canvas.paste(im, ((tw - im.width) // 2, (th - im.height) // 2))
        im = canvas
    elif size:
        im.thumbnail(size, Image.Resampling.LANCZOS)
    dest.parent.mkdir(parents=True, exist_ok=True)
    im.save(dest, "WEBP", quality=82, method=6)
    return im.size


def upload_media(path: Path, alt: str):
    data = path.read_bytes()
    headers = {
        "Content-Disposition": f'attachment; filename="{path.name}"',
        "Content-Type": "image/webp",
    }
    media = api("POST", "media", raw_body=data, headers=headers)
    api("POST", f"media/{media['id']}", {"alt_text": alt, "title": path.stem})
    return media


def ol(items):
    lines = ['<!-- wp:list {"ordered":true} -->', '<ol class="wp-block-list">']
    for item in items:
        lines.append(f"<li>{escape(item)}</li>")
    lines += ["</ol>", "<!-- /wp:list -->"]
    return "\n".join(lines)


def h2(text):
    return f'<!-- wp:heading -->\n<h2 class="wp-block-heading">{escape(text)}</h2>\n<!-- /wp:heading -->'


def h3(text):
    return f'<!-- wp:heading {{"level":3}} -->\n<h3 class="wp-block-heading">{escape(text)}</h3>\n<!-- /wp:heading -->'


def para(text):
    return f'<!-- wp:paragraph -->\n<p>{text}</p>\n<!-- /wp:paragraph -->'


def image_block(mid, src, alt, w, h):
    return (
        f'<!-- wp:image {{"id":{mid},"sizeSlug":"large","linkDestination":"none"}} -->\n'
        f'<figure class="wp-block-image size-large">'
        f'<img src="{src}" alt="{escape(alt)}" class="wp-image-{mid}" width="{w}" height="{h}" loading="lazy" decoding="async"/>'
        f"</figure>\n<!-- /wp:image -->"
    )


def build_carousel(product_media):
    cards = []
    dots = []
    for i, p in enumerate(product_media):
        alt = f"Happy Diwali gift idea: {p['name']} from BlueStone"
        cards.append(
            f'    <div class="bs-cf-card" data-i="{i}">\n'
            f'      <a class="bs-cf-media" href="{p["url"]}">\n'
            f'        <img src="{p["src"]}" alt="{escape(alt)}" width="960" height="535" loading="lazy" decoding="async"/>\n'
            f"      </a>\n"
            f'      <div class="bs-cf-meta">\n'
            f'        <p class="bs-cf-name">{escape(p["name"])}</p>\n'
            f'        <a class="bs-cf-cta" href="{p["url"]}">Buy now</a>\n'
            f"      </div>\n"
            f"    </div>"
        )
        active = " is-active" if i == 0 else ""
        dots.append(
            f'    <button type="button" class="bs-cf-dot{active}" data-i="{i}" aria-label="Product {i+1}"></button>'
        )
    style = Path(ROOT / "output/_eid_carousel_6_snippet.html").read_text().split("<style>")[1].split("</style>")[0]
    script = Path(ROOT / "output/_eid_carousel_6_snippet.html").read_text().split("<script>")[1].split("</script>")[0]
    script = script.replace("bs-cf-eid", "bs-cf-diwali")
    return (
        "<!-- wp:html -->\n<style>\n"
        + style
        + '\n</style>\n<div class="bs-cf" id="bs-cf-diwali" data-interval="3200" aria-roledescription="carousel" aria-label="BlueStone Diwali gift ideas">\n'
        + '  <button type="button" class="bs-cf-nav bs-cf-prev" aria-label="Previous">&#8249;</button>\n'
        + '  <button type="button" class="bs-cf-nav bs-cf-next" aria-label="Next">&#8250;</button>\n'
        + '  <div class="bs-cf-stage">\n'
        + "\n".join(cards)
        + "\n  </div>\n  <div class=\"bs-cf-dots\" role=\"tablist\">\n"
        + "\n".join(dots)
        + "\n  </div>\n</div>\n<script>\n"
        + script
        + "\n</script>\n<!-- /wp:html -->"
    )


def build_faqs():
    faqs = [
        (
            "What is a good short Happy Diwali message for WhatsApp?",
            "Try: \"Happy Diwali 2026! May your home glow with diyas and your heart with peace.\" Short Diwali wishes work best when they feel warm and easy to forward.",
        ),
        (
            "What does adv happy diwali mean?",
            "People often search \"adv happy diwali\" when they want advance Diwali wishes to send before the festival. Use early greetings to wish colleagues, neighbours, or family ahead of the main celebration day.",
        ),
        (
            "What are traditional Diwali wishes for family?",
            "Traditional messages mention light, prosperity, and blessings: \"May Goddess Lakshmi bless your home with wealth, health, and harmony this Diwali.\" Keep the tone respectful and sincere.",
        ),
        (
            "What are good Diwali quotes for Instagram captions?",
            "Short captions perform well: \"Diya mode: on. Gratitude mode: also on.\" Pair a one-line caption with a clear festive photo rather than long text on the image.",
        ),
        (
            "How do I wish someone a happy Diwali in English?",
            "Use clear festival language: \"Wishing you a joyous and prosperous Diwali\" or \"Shubh Deepavali 2026.\" Match formality to your relationship: warm for friends, slightly formal for colleagues.",
        ),
        (
            "What are religious Diwali blessings I can share?",
            "Religious Diwali messages often invoke Lakshmi, Ganesha, or Rama: \"May Lord Ganesha remove obstacles from your path this Diwali and always.\" Share these when you know the recipient appreciates spiritual greetings.",
        ),
        (
            "When should I send Diwali wishes in 2026?",
            "Send advance wishes a few days before Diwali and main greetings on the festival day. Early messages help busy contacts feel remembered without last-minute clutter.",
        ),
    ]
    html = [h2("Frequently Asked Questions about Happy Diwali Wishes")]
    schema = []
    for q, a in faqs:
        html.append(h3(q))
        html.append(para(escape(a)))
        schema.append(
            {
                "@type": "Question",
                "name": q,
                "acceptedAnswer": {"@type": "Answer", "text": a},
            }
        )
    return "\n\n".join(html), schema


def build_content(flatlay, lifestyle, carousel):
    faq_html, faq_schema = build_faqs()
    parts = [
        '<!-- wp:html -->\n<style>\n.bs-eeat{margin:0 auto 1.25rem;max-width:720px;text-align:center;font-size:.95rem;color:#444;line-height:1.5}\n.bs-eeat strong{color:#111}\n.entry-content img,.wp-block-image img{max-width:100%;height:auto}\n</style>\n<!-- /wp:html -->',
        '<!-- wp:paragraph {"align":"center"} -->\n<p class="has-text-align-center bs-eeat">By <strong>Vikas</strong>, BlueStone Editorial</p>\n<!-- /wp:paragraph -->',
        para(
            "Looking for Happy Diwali wishes you can copy, edit, and send today? "
            "Here are traditional greetings, heartfelt quotes, short WhatsApp lines, and captions "
            "for family, friends, and colleagues, all ready for Diwali 2026."
        ),
        para(
            "<strong>TL;DR:</strong> Pick a one-line WhatsApp wish for a quick send, "
            "a traditional blessing for family elders, and a short caption for your festive photo. "
            "Scroll the lists below and use what sounds like you."
        ),
        para(
            "Diwali is the festival of lights, a time when homes glow with diyas, tables fill with sweets, "
            "and messages travel faster than fireworks. Whether your search is <em>adv happy diwali</em> "
            "because you want to greet someone early, or you need the perfect line for a card, "
            "these wishes are grouped so you can find the right tone fast."
        ),
        h2("Short Happy Diwali Messages for WhatsApp"),
        para("Use these when you need a quick, warm happy diwali msg that fits on one screen."),
        ol(SECTIONS["whatsapp"]),
        h2("Traditional Diwali Wishes"),
        para("Classic Shubh Deepavali greetings that work for cards, elders, and formal notes."),
        ol(SECTIONS["traditional"]),
        h2("Heartfelt Diwali Quotes"),
        para("Emotional Diwali messages for people who mean more than a generic greeting."),
        ol(SECTIONS["heartfelt"]),
        image_block(flatlay["id"], flatlay["src"], flatlay["alt"], flatlay["w"], flatlay["h"]),
        h2("Inspirational Diwali Quotes"),
        para("Uplifting lines that fit speeches, office notes, or reflection posts during the festival."),
        ol(SECTIONS["inspirational"]),
        h2("A soft Diwali gift idea (if you are pairing wishes with jewellery)"),
        para(
            "If you are gifting too, let the message lead and the jewellery follow. "
            "Festive pieces that work beyond one night often feel more thoughtful. "
            "Browse these BlueStone picks if you want a starting point for Diwali 2026."
        ),
        carousel,
        h2("Lighthearted Diwali Greetings"),
        para("Funny and playful Diwali wishes for friends who appreciate humour with their mithai."),
        ol(SECTIONS["lighthearted"]),
        h2("Short and Sweet Diwali Wishes"),
        para("Ultra-short happy diwali greetings for status updates, tags, and quick replies."),
        ol(SECTIONS["short"]),
        h2("Religious Diwali Blessings"),
        para("Spiritual Diwali messages that honour faith, devotion, and traditional blessings."),
        ol(SECTIONS["religious"]),
        image_block(lifestyle["id"], lifestyle["src"], lifestyle["alt"], lifestyle["w"], lifestyle["h"]),
        h2("Diwali Wishes for Family and Friends"),
        para("Warm messages for the people you celebrate with, near or far."),
        ol(SECTIONS["family"]),
        h2("Diwali Captions for Instagram & Status"),
        para("Short captions for festive photos, stories, and WhatsApp status."),
        ol(SECTIONS["captions"]),
        h2("Unique Diwali Greetings"),
        para("Stand-out lines when you want something less common than the usual festival text."),
        ol(SECTIONS["unique"]),
        h2("How to pick the right Diwali wish in 10 seconds"),
        para("Match tone to relationship: elders appreciate traditional blessings; friends may prefer lighthearted lines."),
        para("Match length to medium: WhatsApp favours one sentence; cards can hold a short paragraph."),
        para("Send advance wishes early if your calendar is packed on the main day."),
        para("Say it in your voice. If you would not normally use a phrase, skip it."),
        h2("More festive reads to explore"),
        para(
            'Planning other moments too? Save our '
            '<a href="https://blog.bluestone.com/happy-holi-wishes-messages-quotes-2027/">Happy Holi wishes for 2027</a>, '
            '<a href="https://blog.bluestone.com/happy-eid-mubarak-wishes-messages-quotes-2027/">Eid Mubarak wishes for 2027</a>, '
            '<a href="https://blog.bluestone.com/best-dhanteras-wishes-messages-greetings/">Dhanteras wishes</a>, and '
            '<a href="https://blog.bluestone.com/heart-touching-raksha-bandhan-wishes-quotes/">Raksha Bandhan wishes</a> '
            'for copy-ready messages across the year.'
        ),
        para(
            'For context on Diwali traditions, see '
            '<a href="https://en.wikipedia.org/wiki/Diwali" rel="noopener">Diwali on Wikipedia</a>.'
        ),
        h2("Conclusion"),
        para(
            "The best Happy Diwali wishes sound like you on a generous day. Choose words that match your relationship, "
            "send them through the medium that feels natural, and let the festival do the rest. "
            "Whether you are sharing advance greetings or lighting the first diya, sincerity beats performance every time."
        ),
        faq_html,
    ]
    return "\n\n".join(parts), faq_schema


def update_xlsx(post_id, url):
    import openpyxl

    wb = openpyxl.load_workbook(ROOT / "SEO Strategy 2026.xlsx")
    ws = wb["Week 1-2"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    for row in ws.iter_rows(min_row=2):
        if row[idx["Rank"]].value == 4:
            row[idx["Suggested URL Slug"]].value = SLUG
            row[idx["Bluestone Blog URL"]].value = url
            note = row[idx["Execution Note"]].value or ""
            base = note.split("| Published")[0].strip() if note else ""
            row[idx["Execution Note"]].value = (
                f"{base} | Published {datetime.now().strftime('%Y-%m-%d')} WP#{post_id} | "
                f"Diwali 2026; author Vikas; Magnific WebP hero + flatlay + lifestyle; 6-product coverflow carousel"
            )
            break
    wb.save(ROOT / "SEO Strategy 2026.xlsx")


def main():
    for t in TYPE3:
        if not t["src"].exists():
            raise SystemExit(f"Missing Type 3 image: {t['src']}. Run magnific_generate_images.py --preset diwali-2026")

    assets = ROOT / "output/Week1_Rank4_Diwali_assets"
    assets.mkdir(exist_ok=True)

    product_media = []
    for p in PRODUCTS:
        if not p["png"].exists():
            raise SystemExit(f"Missing product PNG: {p['png']}")
        webp = assets / (p["name"].replace(" ", "-") + "-carousel.webp")
        to_webp(p["png"], webp, carousel=True)
        alt = f"Happy Diwali gift idea: {p['name']} from BlueStone"
        media = upload_media(webp, alt)
        product_media.append(
            {
                "name": p["name"],
                "url": p["url"],
                "id": media["id"],
                "src": media["source_url"],
            }
        )
        print("product", p["name"], media["id"])

    concept = {}
    hero_media = None
    for t in TYPE3:
        webp = assets / t["filename"]
        if t["src"] == webp:
            w, h = to_webp(t["src"], webp, size=(1400, 933))
        else:
            import shutil
            shutil.copy2(t["src"], webp)
            from PIL import Image
            im = Image.open(webp)
            w, h = im.size
        media = upload_media(webp, t["alt"])
        info = {"id": media["id"], "src": media["source_url"], "alt": t["alt"], "w": w, "h": h}
        concept[t["key"]] = info
        if t["featured"]:
            hero_media = media["id"]
        print("type3", t["key"], media["id"])

    carousel = build_carousel(product_media)
    content, faq_schema = build_content(concept["flatlay"], concept["lifestyle"], carousel)

    images = [concept["hero"]["src"], concept["flatlay"]["src"], concept["lifestyle"]["src"]] + [p["src"] for p in product_media]
    article_schema = {
        "@context": "https://schema.org",
        "@type": "BlogPosting",
        "headline": TITLE,
        "description": META_DESC,
        "datePublished": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "dateModified": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "author": {"@type": "Person", "name": "Vikas"},
        "publisher": {"@type": "Organization", "name": "BlueStone", "url": "https://www.bluestone.com/"},
        "image": images,
        "mainEntityOfPage": {
            "@type": "WebPage",
            "@id": f"https://blog.bluestone.com/{SLUG}/",
        },
        "keywords": [
            "adv happy diwali",
            "happy diwali wishes",
            "diwali messages",
            "diwali quotes",
            "shubh deepavali",
        ],
    }
    faq_page = {"@context": "https://schema.org", "@type": "FAQPage", "mainEntity": faq_schema}
    content += (
        "\n\n<!-- wp:html -->\n"
        f'<script type="application/ld+json">{json.dumps(faq_page, ensure_ascii=False)}</script>\n'
        f'<script type="application/ld+json">{json.dumps(article_schema, ensure_ascii=False)}</script>\n'
        "<!-- /wp:html -->\n"
    )

    post = api(
        "POST",
        "posts",
        {
            "title": TITLE,
            "slug": SLUG,
            "status": "publish",
            "author": 270271338,
            "featured_media": hero_media,
            "content": content,
            "excerpt": META_DESC,
            "meta": {
                "_yoast_wpseo_focuskw": FOCUS_KW,
                "_yoast_wpseo_title": YOAST_TITLE,
                "_yoast_wpseo_metadesc": META_DESC,
            },
        },
    )
    print("published", post["id"], post["link"])

    Path(ROOT / "output/Week1_Rank4_Diwali_product_media.json").write_text(
        json.dumps(product_media, indent=2)
    )
    update_xlsx(post["id"], post["link"])

    html = urllib.request.urlopen(
        urllib.request.Request(post["link"] + "?v=audit", headers={"User-Agent": "Mozilla/5.0"}),
        timeout=30,
    ).read().decode("utf-8", "replace")
    checks = {
        "H1": len(re.findall(r"<h1\b", html, re.I)) == 1,
        "FAQ": "Frequently Asked Questions" in html,
        "carousel": "bs-cf-diwali" in html,
        "buy now": html.count("Buy now") >= 6,
        "Vikas": "Vikas" in html,
        "2026": "2026" in html,
    }
    print("audit", checks)


if __name__ == "__main__":
    main()

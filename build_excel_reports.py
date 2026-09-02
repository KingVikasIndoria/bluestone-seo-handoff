#!/usr/bin/env python3
"""
Generate Excel Reports for Google Search Console and SEO Dashboard
=====================================================================
1. Google_Search_Console_Report.xlsx
   - Clicks, Impressions, CTR, Average Position, Index Coverage
2. SEO_Dashboard_Report.xlsx
   - Published Blogs, Indexed Blogs, Clicks, Impressions, CTR, Ranking Distribution
"""

import json
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load data
BASE_DIR = Path(__file__).parent
DATA_PATH = BASE_DIR / "dashboard_data.json"

with open(DATA_PATH, "r", encoding="utf-8") as f:
    data = json.load(f)

metadata = data.get("metadata", {})
strat_cmp = metadata.get("strategy_comparison", {})
rank_3w = metadata.get("rank_breakdown_3w", {})
all_blogs = data.get("all_blogs", [])
post_july16_blogs = data.get("post_july16_blogs", [])
pre_july16_blogs = data.get("pre_july16_blogs", [])
striking_blogs = data.get("striking_distance", [])
top100_blogs = data.get("top_100_performing", [])
weekly_trends = data.get("weekly_trends_10w", [])
monthly_6m = data.get("monthly_trends_6m", [])
monthly_12m = data.get("monthly_trends_12m", [])
indexing_trend = data.get("post_july16_indexing_trend", [])
devices_data = data.get("devices", [])

# Fonts & Colors
FONT_NAME = "Calibri"

title_font = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
subtitle_font = Font(name=FONT_NAME, size=11, italic=True, color="E2E8F0")
section_font = Font(name=FONT_NAME, size=12, bold=True, color="1E293B")
header_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
subhdr_font = Font(name=FONT_NAME, size=10, bold=True, color="1E293B")
data_font = Font(name=FONT_NAME, size=10, color="0F172A")
bold_data_font = Font(name=FONT_NAME, size=10, bold=True, color="0F172A")

kpi_val_font = Font(name=FONT_NAME, size=18, bold=True, color="1E3A8A")
kpi_lbl_font = Font(name=FONT_NAME, size=9, bold=True, color="475569")

# Fills
title_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy Blue
header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid") # Royal Blue
subhdr_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Soft Blue-Gray
kpi_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

badge_indexed_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light Green
badge_indexed_font = Font(name=FONT_NAME, size=10, color="166534", bold=True)

badge_pending_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Amber
badge_pending_font = Font(name=FONT_NAME, size=10, color="92400E", bold=True)

# Borders
thin_side = Side(border_style="thin", color="CBD5E1")
thick_bottom_side = Side(border_style="medium", color="1E3A8A")
double_bottom_side = Side(border_style="double", color="1E3A8A")

border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom_side)
border_total = Border(left=thin_side, right=thin_side, top=thin_side, bottom=double_bottom_side)

# Alignments
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_title_banner(ws, title, subtitle, max_col=10):
    col_let = get_column_letter(max_col)
    ws.merge_cells(f"A1:{col_let}1")
    ws.merge_cells(f"A2:{col_let}2")
    
    cell1 = ws["A1"]
    cell1.value = title
    cell1.font = title_font
    cell1.fill = title_fill
    cell1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    cell2 = ws["A2"]
    cell2.value = subtitle
    cell2.font = subtitle_font
    cell2.fill = title_fill
    cell2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

def create_kpi_card(ws, start_row, start_col, label, value, num_format=None):
    c1 = ws.cell(row=start_row, column=start_col)
    c1.value = label.upper()
    c1.font = kpi_lbl_font
    c1.fill = kpi_fill
    c1.alignment = Alignment(horizontal="center", vertical="center")
    
    c2 = ws.cell(row=start_row+1, column=start_col)
    c2.value = value
    c2.font = kpi_val_font
    c2.fill = kpi_fill
    c2.alignment = Alignment(horizontal="center", vertical="center")
    if num_format:
        c2.number_format = num_format
        
    for r in range(start_row, start_row+2):
        cell = ws.cell(row=r, column=start_col)
        cell.border = border_cell

def style_header_row(ws, row_idx, headers):
    ws.row_dimensions[row_idx].height = 26
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_header
        cell.border = border_header

def autofit_columns(ws, max_cols=20):
    ws.views.sheetView[0].showGridLines = True
    for col in range(1, max_cols + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is not None:
                s_val = str(cell_val)
                if row in (1, 2) and col == 1:
                    continue
                if len(s_val) > max_len:
                    max_len = len(s_val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 65)

# ==============================================================================
# REPORT 1: GOOGLE SEARCH CONSOLE REPORT
# ==============================================================================
wb_gsc = openpyxl.Workbook()

# Sheet 1: GSC Summary & KPIs
ws_gsc_sum = wb_gsc.active
ws_gsc_sum.title = "GSC Summary & KPIs"

apply_title_banner(
    ws_gsc_sum,
    "Google Search Console Performance Report",
    f"Domain: sc-domain:bluestone.com | Target: blog.bluestone.com | Period: {metadata.get('start_date')} to {metadata.get('end_date')} | Generated: {metadata.get('generated_at')}",
    max_col=8
)

# KPI Cards
create_kpi_card(ws_gsc_sum, 4, 1, "Total Clicks", strat_cmp.get("overall", {}).get("total_clicks", 0), "#,##0")
create_kpi_card(ws_gsc_sum, 4, 2, "Total Impressions", strat_cmp.get("overall", {}).get("total_impressions", 0), "#,##0")
create_kpi_card(ws_gsc_sum, 4, 3, "Average CTR", strat_cmp.get("overall", {}).get("avg_ctr", 0) / 100.0, "0.00%")
create_kpi_card(ws_gsc_sum, 4, 4, "Average Position", strat_cmp.get("overall", {}).get("avg_position", 0), "0.0")
create_kpi_card(ws_gsc_sum, 4, 5, "Total Published URLs", strat_cmp.get("overall", {}).get("count", 0), "#,##0")
create_kpi_card(ws_gsc_sum, 4, 6, "Total Indexed URLs", strat_cmp.get("overall", {}).get("indexed_count", 0), "#,##0")
create_kpi_card(ws_gsc_sum, 4, 7, "Index Coverage Rate", strat_cmp.get("overall", {}).get("indexing_rate", 0) / 100.0, "0.0%")
create_kpi_card(ws_gsc_sum, 4, 8, "Pending Index URLs", strat_cmp.get("overall", {}).get("count", 0) - strat_cmp.get("overall", {}).get("indexed_count", 0), "#,##0")

# Section 1: Strategy Segment Performance Comparison
ws_gsc_sum.cell(row=7, column=1, value="Search Console Metrics by Strategy Segment").font = section_font

gsc_sum_headers = [
    "Strategy Segment", "Total Blogs", "Indexed Blogs", "Indexing Rate (%)",
    "Total Clicks", "Total Impressions", "Average CTR (%)", "Average Position"
]
style_header_row(ws_gsc_sum, 8, gsc_sum_headers)

seg_rows = [
    ("🚀 New Strategy (Post-July 16)", strat_cmp.get("new_strategy", {})),
    ("📜 Legacy Strategy (Pre-July 16)", strat_cmp.get("legacy_strategy", {})),
    ("📊 Total / Overall Average", strat_cmp.get("overall", {}))
]

for idx, (label, sdata) in enumerate(seg_rows, start=9):
    ws_gsc_sum.row_dimensions[idx].height = 22
    is_total = (idx == 11)
    
    r_font = bold_data_font if is_total else data_font
    r_fill = total_fill if is_total else PatternFill(fill_type=None)
    r_border = border_total if is_total else border_cell
    
    c_label = ws_gsc_sum.cell(row=idx, column=1, value=label)
    c_count = ws_gsc_sum.cell(row=idx, column=2, value=sdata.get("count", 0))
    c_indexed = ws_gsc_sum.cell(row=idx, column=3, value=sdata.get("indexed_count", 0))
    c_rate = ws_gsc_sum.cell(row=idx, column=4, value=sdata.get("indexing_rate", 0) / 100.0)
    c_clicks = ws_gsc_sum.cell(row=idx, column=5, value=sdata.get("total_clicks", 0))
    c_imp = ws_gsc_sum.cell(row=idx, column=6, value=sdata.get("total_impressions", 0))
    c_ctr = ws_gsc_sum.cell(row=idx, column=7, value=sdata.get("avg_ctr", 0) / 100.0)
    c_pos = ws_gsc_sum.cell(row=idx, column=8, value=sdata.get("avg_position", 0))
    
    c_count.number_format = "#,##0"
    c_indexed.number_format = "#,##0"
    c_rate.number_format = "0.0%"
    c_clicks.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    c_pos.number_format = "0.0"
    
    for c_i, c_cell in enumerate([c_label, c_count, c_indexed, c_rate, c_clicks, c_imp, c_ctr, c_pos], 1):
        c_cell.font = r_font
        if r_fill.fill_type:
            c_cell.fill = r_fill
        c_cell.border = r_border
        c_cell.alignment = align_left if c_i == 1 else align_right

# Section 2: Device Performance Breakdown
ws_gsc_sum.cell(row=13, column=1, value="Performance Breakdown by Device Category").font = section_font

device_headers = ["Device Category", "Total Clicks", "Total Impressions", "CTR (%)", "Average Position", "% Share of Total Clicks"]
style_header_row(ws_gsc_sum, 14, device_headers)

total_dev_clicks = sum(d.get("clicks", 0) for d in devices_data) or 1
for idx, d in enumerate(devices_data, start=15):
    ws_gsc_sum.row_dimensions[idx].height = 20
    
    dev_name = d.get("device", "").upper()
    dev_clicks = d.get("clicks", 0)
    dev_imp = d.get("impressions", 0)
    dev_ctr = d.get("ctr", 0) / 100.0
    dev_pos = d.get("position", 0)
    dev_share = dev_clicks / total_dev_clicks
    
    c_dev = ws_gsc_sum.cell(row=idx, column=1, value=dev_name)
    c_clk = ws_gsc_sum.cell(row=idx, column=2, value=dev_clicks)
    c_imp = ws_gsc_sum.cell(row=idx, column=3, value=dev_imp)
    c_ctr = ws_gsc_sum.cell(row=idx, column=4, value=dev_ctr)
    c_pos = ws_gsc_sum.cell(row=idx, column=5, value=dev_pos)
    c_shr = ws_gsc_sum.cell(row=idx, column=6, value=dev_share)
    
    c_clk.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    c_pos.number_format = "0.0"
    c_shr.number_format = "0.0%"
    
    for c_i, c_cell in enumerate([c_dev, c_clk, c_imp, c_ctr, c_pos, c_shr], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        c_cell.alignment = align_left if c_i == 1 else align_right

autofit_columns(ws_gsc_sum, max_cols=8)

# ------------------------------------------------------------------------------
# Sheet 2: Index Coverage Report
# ------------------------------------------------------------------------------
ws_gsc_idx = wb_gsc.create_sheet(title="Index Coverage Report")

apply_title_banner(
    ws_gsc_idx,
    "Google Index Coverage & Discovery Analysis",
    "Index status breakdown across WordPress blog portfolio",
    max_col=10
)

ws_gsc_idx.cell(row=4, column=1, value="Index Status Overview").font = section_font
idx_ov_headers = ["Indexing Status", "Published Blogs", "% Share of Total", "Indexing Velocity / Lag Notes"]
style_header_row(ws_gsc_idx, 5, idx_ov_headers)

tot_count = strat_cmp.get("overall", {}).get("count", 1564)
ind_count = strat_cmp.get("overall", {}).get("indexed_count", 1322)
pend_count = tot_count - ind_count

idx_ov_rows = [
    ("Indexed (Active in Search)", ind_count, ind_count / tot_count, "Successfully indexed by Googlebot"),
    ("Unindexed / Pending Inspection", pend_count, pend_count / tot_count, "Submitted via Indexing API / Pending crawl"),
    ("Total Published Portfolio", tot_count, 1.0, "Complete WordPress blog inventory")
]

for idx, (status, count, pct, note) in enumerate(idx_ov_rows, start=6):
    ws_gsc_idx.row_dimensions[idx].height = 20
    is_tot = (idx == 8)
    
    c_st = ws_gsc_idx.cell(row=idx, column=1, value=status)
    c_ct = ws_gsc_idx.cell(row=idx, column=2, value=count)
    c_pc = ws_gsc_idx.cell(row=idx, column=3, value=pct)
    c_nt = ws_gsc_idx.cell(row=idx, column=4, value=note)
    
    c_ct.number_format = "#,##0"
    c_pc.number_format = "0.0%"
    
    r_font = bold_data_font if is_tot else data_font
    r_border = border_total if is_tot else border_cell
    
    for c_i, c_cell in enumerate([c_st, c_ct, c_pc, c_nt], 1):
        c_cell.font = r_font
        c_cell.border = r_border
        if is_tot:
            c_cell.fill = total_fill
        c_cell.alignment = align_left if c_i in (1, 4) else align_right

# Master List of Indexing Status per Blog URL
ws_gsc_idx.cell(row=10, column=1, value="Complete Portfolio Indexing Audit (1,564 URLs)").font = section_font

idx_master_headers = [
    "WP ID", "Blog Title", "URL", "Published Date", "Strategy Segment",
    "Indexing Status", "First Indexed Date", "Indexing Lag (Days)", "Clicks (30d)", "Impressions (30d)"
]
style_header_row(ws_gsc_idx, 11, idx_master_headers)

for idx, b in enumerate(all_blogs, start=12):
    ws_gsc_idx.row_dimensions[idx].height = 19
    
    is_ind = b.get("is_indexed", False)
    st_badge = "Indexed" if is_ind else "Pending Index"
    
    c_id = ws_gsc_idx.cell(row=idx, column=1, value=b.get("wp_id"))
    c_ttl = ws_gsc_idx.cell(row=idx, column=2, value=b.get("title"))
    c_url = ws_gsc_idx.cell(row=idx, column=3, value=b.get("link"))
    c_pub = ws_gsc_idx.cell(row=idx, column=4, value=b.get("published_date"))
    c_str = ws_gsc_idx.cell(row=idx, column=5, value="New Strategy" if b.get("is_new_strategy") else "Legacy Strategy")
    c_idx = ws_gsc_idx.cell(row=idx, column=6, value=st_badge)
    c_fidx = ws_gsc_idx.cell(row=idx, column=7, value=b.get("first_indexed_date", "Not Indexed Yet"))
    c_lag = ws_gsc_idx.cell(row=idx, column=8, value=b.get("indexing_lag_days"))
    c_clk = ws_gsc_idx.cell(row=idx, column=9, value=b.get("clicks", 0))
    c_imp = ws_gsc_idx.cell(row=idx, column=10, value=b.get("impressions", 0))
    
    c_id.number_format = "0"
    c_clk.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    if b.get("indexing_lag_days") is not None:
        c_lag.number_format = "#,##0"
        
    for c_i, c_cell in enumerate([c_id, c_ttl, c_url, c_pub, c_str, c_idx, c_fidx, c_lag, c_clk, c_imp], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        if c_i in (1, 4, 5, 7):
            c_cell.alignment = align_center
        elif c_i in (2, 3):
            c_cell.alignment = align_left
        else:
            c_cell.alignment = align_right
            
    # Highlight Indexing badge
    if is_ind:
        c_idx.fill = badge_indexed_fill
        c_idx.font = badge_indexed_font
    else:
        c_idx.fill = badge_pending_fill
        c_idx.font = badge_pending_font
        
autofit_columns(ws_gsc_idx, max_cols=10)

# ------------------------------------------------------------------------------
# Sheet 3: Page Performance (GSC)
# ------------------------------------------------------------------------------
ws_gsc_page = wb_gsc.create_sheet(title="Page Performance (GSC)")

apply_title_banner(
    ws_gsc_page,
    "Search Console Page Performance Master Export",
    "Detailed page-level metrics (Clicks, Impressions, CTR, Position) for all published blogs",
    max_col=12
)

page_headers = [
    "WP ID", "Blog Title", "URL", "Strategy Segment", "Published Date",
    "Indexing Status", "Clicks (30d)", "Impressions (30d)", "CTR (%)", "Average Position",
    "Top Target Query", "Health Status"
]
style_header_row(ws_gsc_page, 4, page_headers)

sorted_blogs = sorted(all_blogs, key=lambda x: (x.get("clicks", 0), x.get("impressions", 0)), reverse=True)

for idx, b in enumerate(sorted_blogs, start=5):
    ws_gsc_page.row_dimensions[idx].height = 19
    
    is_ind = b.get("is_indexed", False)
    st_badge = "Indexed" if is_ind else "Pending Index"
    
    top_q = b.get("top_queries", [])
    primary_q = top_q[0]["query"] if top_q else "N/A"
    
    c_id = ws_gsc_page.cell(row=idx, column=1, value=b.get("wp_id"))
    c_ttl = ws_gsc_page.cell(row=idx, column=2, value=b.get("title"))
    c_url = ws_gsc_page.cell(row=idx, column=3, value=b.get("link"))
    c_str = ws_gsc_page.cell(row=idx, column=4, value="New Strategy" if b.get("is_new_strategy") else "Legacy Strategy")
    c_pub = ws_gsc_page.cell(row=idx, column=5, value=b.get("published_date"))
    c_idx = ws_gsc_page.cell(row=idx, column=6, value=st_badge)
    c_clk = ws_gsc_page.cell(row=idx, column=7, value=b.get("clicks", 0))
    c_imp = ws_gsc_page.cell(row=idx, column=8, value=b.get("impressions", 0))
    c_ctr = ws_gsc_page.cell(row=idx, column=9, value=b.get("ctr", 0) / 100.0)
    c_pos = ws_gsc_page.cell(row=idx, column=10, value=b.get("position", 0))
    c_qry = ws_gsc_page.cell(row=idx, column=11, value=primary_q)
    c_hlt = ws_gsc_page.cell(row=idx, column=12, value=b.get("health"))
    
    c_id.number_format = "0"
    c_clk.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    c_pos.number_format = "0.0"
    
    for c_i, c_cell in enumerate([c_id, c_ttl, c_url, c_str, c_pub, c_idx, c_clk, c_imp, c_ctr, c_pos, c_qry, c_hlt], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        if c_i in (1, 4, 5, 6):
            c_cell.alignment = align_center
        elif c_i in (2, 3, 11, 12):
            c_cell.alignment = align_left
        else:
            c_cell.alignment = align_right
            
    if is_ind:
        c_idx.fill = badge_indexed_fill
        c_idx.font = badge_indexed_font
    else:
        c_idx.fill = badge_pending_fill
        c_idx.font = badge_pending_font

autofit_columns(ws_gsc_page, max_cols=12)

# ------------------------------------------------------------------------------
# Sheet 4: Search Queries Report
# ------------------------------------------------------------------------------
ws_gsc_query = wb_gsc.create_sheet(title="Search Queries")

apply_title_banner(
    ws_gsc_query,
    "Google Search Console Queries Master Export",
    "Query-level search analytics mapped to landing blog URLs",
    max_col=8
)

query_headers = [
    "Primary Search Query", "Target Blog Title", "Landing URL", "Query Clicks",
    "Query Impressions", "Query CTR (%)", "Query Average Position", "Strategy Segment"
]
style_header_row(ws_gsc_query, 4, query_headers)

query_rows = []
for b in all_blogs:
    t_queries = b.get("top_queries", [])
    for q in t_queries:
        query_rows.append({
            "query": q.get("query"),
            "title": b.get("title"),
            "url": b.get("link"),
            "clicks": q.get("clicks", 0),
            "impressions": q.get("impressions", 0),
            "ctr": q.get("ctr", 0) / 100.0,
            "position": q.get("position", 0),
            "strategy": "New Strategy" if b.get("is_new_strategy") else "Legacy Strategy"
        })

# Sort queries by impressions/clicks
query_rows.sort(key=lambda x: (x["clicks"], x["impressions"]), reverse=True)

for idx, q_item in enumerate(query_rows[:5000], start=5):
    ws_gsc_query.row_dimensions[idx].height = 19
    
    c_q = ws_gsc_query.cell(row=idx, column=1, value=q_item["query"])
    c_t = ws_gsc_query.cell(row=idx, column=2, value=q_item["title"])
    c_u = ws_gsc_query.cell(row=idx, column=3, value=q_item["url"])
    c_c = ws_gsc_query.cell(row=idx, column=4, value=q_item["clicks"])
    c_i = ws_gsc_query.cell(row=idx, column=5, value=q_item["impressions"])
    c_ctr = ws_gsc_query.cell(row=idx, column=6, value=q_item["ctr"])
    c_p = ws_gsc_query.cell(row=idx, column=7, value=q_item["position"])
    c_s = ws_gsc_query.cell(row=idx, column=8, value=q_item["strategy"])
    
    c_c.number_format = "#,##0"
    c_i.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    c_p.number_format = "0.0"
    
    for c_idx_i, c_cell in enumerate([c_q, c_t, c_u, c_c, c_i, c_ctr, c_p, c_s], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        if c_idx_i in (1, 2, 3):
            c_cell.alignment = align_left
        elif c_idx_i == 8:
            c_cell.alignment = align_center
        else:
            c_cell.alignment = align_right

autofit_columns(ws_gsc_query, max_cols=8)

# ------------------------------------------------------------------------------
# Sheet 5: Performance Trends
# ------------------------------------------------------------------------------
ws_gsc_trend = wb_gsc.create_sheet(title="Performance Trends")

apply_title_banner(
    ws_gsc_trend,
    "Google Search Console Historical Performance Trends",
    "Weekly and monthly traffic progression (Clicks & Impressions)",
    max_col=6
)

# 10-Week Trend Table
ws_gsc_trend.cell(row=4, column=1, value="10-Week Weekly Performance Trend").font = section_font
w_headers = ["Week Period", "Clicks", "Impressions", "Calculated CTR (%)"]
style_header_row(ws_gsc_trend, 5, w_headers)

for idx, w in enumerate(weekly_trends, start=6):
    ws_gsc_trend.row_dimensions[idx].height = 20
    w_clk = w.get("clicks", 0)
    w_imp = w.get("impressions", 0)
    w_ctr = (w_clk / w_imp) if w_imp > 0 else 0
    
    c_lbl = ws_gsc_trend.cell(row=idx, column=1, value=w.get("week_label"))
    c_clk = ws_gsc_trend.cell(row=idx, column=2, value=w_clk)
    c_imp = ws_gsc_trend.cell(row=idx, column=3, value=w_imp)
    c_ctr = ws_gsc_trend.cell(row=idx, column=4, value=w_ctr)
    
    c_clk.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    
    for c_i, c_cell in enumerate([c_lbl, c_clk, c_imp, c_ctr], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        c_cell.alignment = align_left if c_i == 1 else align_right

# 6-Month Trend Table
start_r_6m = 6 + len(weekly_trends) + 2
ws_gsc_trend.cell(row=start_r_6m, column=1, value="6-Month Monthly Performance Trend").font = section_font
style_header_row(ws_gsc_trend, start_r_6m + 1, ["Month Label", "Clicks", "Impressions", "Calculated CTR (%)"])

for idx, m in enumerate(monthly_6m, start=start_r_6m + 2):
    ws_gsc_trend.row_dimensions[idx].height = 20
    m_clk = m.get("clicks", 0)
    m_imp = m.get("impressions", 0)
    m_ctr = (m_clk / m_imp) if m_imp > 0 else 0
    
    c_lbl = ws_gsc_trend.cell(row=idx, column=1, value=m.get("month_label"))
    c_clk = ws_gsc_trend.cell(row=idx, column=2, value=m_clk)
    c_imp = ws_gsc_trend.cell(row=idx, column=3, value=m_imp)
    c_ctr = ws_gsc_trend.cell(row=idx, column=4, value=m_ctr)
    
    c_clk.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    
    for c_i, c_cell in enumerate([c_lbl, c_clk, c_imp, c_ctr], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        c_cell.alignment = align_left if c_i == 1 else align_right

autofit_columns(ws_gsc_trend, max_cols=6)

# Save Report 1
gsc_path = BASE_DIR / "Google_Search_Console_Report.xlsx"
wb_gsc.save(gsc_path)
print(f"✅ Generated {gsc_path.name}")


# ==============================================================================
# REPORT 2: SEO DASHBOARD REPORT
# ==============================================================================
wb_dash = openpyxl.Workbook()

# Sheet 1: Dashboard Overview
ws_d_sum = wb_dash.active
ws_d_sum.title = "Dashboard Overview"

apply_title_banner(
    ws_d_sum,
    "Bluestone SEO Performance Dashboard Export",
    f"Executive Report | Generated: {metadata.get('generated_at')} | Strategy Pivot: July 16, 2026",
    max_col=9
)

# KPI Block
create_kpi_card(ws_d_sum, 4, 1, "Published Blogs", strat_cmp.get("overall", {}).get("count", 0), "#,##0")
create_kpi_card(ws_d_sum, 4, 2, "Indexed Blogs", strat_cmp.get("overall", {}).get("indexed_count", 0), "#,##0")
create_kpi_card(ws_d_sum, 4, 3, "Indexing Rate", strat_cmp.get("overall", {}).get("indexing_rate", 0) / 100.0, "0.0%")
create_kpi_card(ws_d_sum, 4, 4, "Total Clicks (30d)", strat_cmp.get("overall", {}).get("total_clicks", 0), "#,##0")
create_kpi_card(ws_d_sum, 4, 5, "Total Impressions", strat_cmp.get("overall", {}).get("total_impressions", 0), "#,##0")
create_kpi_card(ws_d_sum, 4, 6, "Average CTR", strat_cmp.get("overall", {}).get("avg_ctr", 0) / 100.0, "0.00%")
create_kpi_card(ws_d_sum, 4, 7, "Average Position", strat_cmp.get("overall", {}).get("avg_position", 0), "0.0")

# Section 1: Published & Indexed Breakdown by Strategy
ws_d_sum.cell(row=7, column=1, value="SEO Dashboard Summary by Content Strategy").font = section_font

d_sum_headers = [
    "Content Strategy Segment", "Published Blogs", "Indexed Blogs", "Indexing Rate (%)",
    "Total Clicks", "Total Impressions", "Average CTR (%)", "Average Position", "Publishing Velocity (30d)"
]
style_header_row(ws_d_sum, 8, d_sum_headers)

d_seg_rows = [
    ("🚀 New Strategy (Post-July 16)", strat_cmp.get("new_strategy", {}), strat_cmp.get("published_last_30_days", 292)),
    ("📜 Legacy Strategy (Pre-July 16)", strat_cmp.get("legacy_strategy", {}), 0),
    ("📊 Total Portfolio", strat_cmp.get("overall", {}), strat_cmp.get("published_last_30_days", 292))
]

for idx, (lbl, sdata, pub_30d) in enumerate(d_seg_rows, start=9):
    ws_d_sum.row_dimensions[idx].height = 22
    is_tot = (idx == 11)
    
    r_font = bold_data_font if is_tot else data_font
    r_fill = total_fill if is_tot else PatternFill(fill_type=None)
    r_border = border_total if is_tot else border_cell
    
    c_lbl = ws_d_sum.cell(row=idx, column=1, value=lbl)
    c_pub = ws_d_sum.cell(row=idx, column=2, value=sdata.get("count", 0))
    c_ind = ws_d_sum.cell(row=idx, column=3, value=sdata.get("indexed_count", 0))
    c_rate = ws_d_sum.cell(row=idx, column=4, value=sdata.get("indexing_rate", 0) / 100.0)
    c_clk = ws_d_sum.cell(row=idx, column=5, value=sdata.get("total_clicks", 0))
    c_imp = ws_d_sum.cell(row=idx, column=6, value=sdata.get("total_impressions", 0))
    c_ctr = ws_d_sum.cell(row=idx, column=7, value=sdata.get("avg_ctr", 0) / 100.0)
    c_pos = ws_d_sum.cell(row=idx, column=8, value=sdata.get("avg_position", 0))
    c_v30 = ws_d_sum.cell(row=idx, column=9, value=pub_30d)
    
    c_pub.number_format = "#,##0"
    c_ind.number_format = "#,##0"
    c_rate.number_format = "0.0%"
    c_clk.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    c_pos.number_format = "0.0"
    c_v30.number_format = "#,##0"
    
    for c_i, c_cell in enumerate([c_lbl, c_pub, c_ind, c_rate, c_clk, c_imp, c_ctr, c_pos, c_v30], 1):
        c_cell.font = r_font
        if r_fill.fill_type:
            c_cell.fill = r_fill
        c_cell.border = r_border
        c_cell.alignment = align_left if c_i == 1 else align_right

# Section 2: Recent Publishing Activity
ws_d_sum.cell(row=13, column=1, value="Recent Content Publishing Activity").font = section_font
pub_headers = ["Time Period", "Published Count", "% of Portfolio", "Strategy Focus"]
style_header_row(ws_d_sum, 14, pub_headers)

tot_p = strat_cmp.get("overall", {}).get("count", 1564)
pub_act_rows = [
    ("Published This Week", strat_cmp.get("published_this_week", 9), strat_cmp.get("published_this_week", 9)/tot_p, "High Intent Commercial Blogs"),
    ("Published Last Week", strat_cmp.get("published_last_week", 97), strat_cmp.get("published_last_week", 97)/tot_p, "Cluster Scaling & Internal Linking"),
    ("Published Last 30 Days", strat_cmp.get("published_last_30_days", 292), strat_cmp.get("published_last_30_days", 292)/tot_p, "New Strategy Expansion")
]

for idx, (p_lbl, p_cnt, p_pct, p_foc) in enumerate(pub_act_rows, start=15):
    ws_d_sum.row_dimensions[idx].height = 20
    
    c_lbl = ws_d_sum.cell(row=idx, column=1, value=p_lbl)
    c_cnt = ws_d_sum.cell(row=idx, column=2, value=p_cnt)
    c_pct = ws_d_sum.cell(row=idx, column=3, value=p_pct)
    c_foc = ws_d_sum.cell(row=idx, column=4, value=p_foc)
    
    c_cnt.number_format = "#,##0"
    c_pct.number_format = "0.0%"
    
    for c_i, c_cell in enumerate([c_lbl, c_cnt, c_pct, c_foc], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        c_cell.alignment = align_left if c_i in (1, 4) else align_right

autofit_columns(ws_d_sum, max_cols=9)

# ------------------------------------------------------------------------------
# Sheet 2: Ranking Distribution
# ------------------------------------------------------------------------------
ws_d_rank = wb_dash.create_sheet(title="Ranking Distribution")

apply_title_banner(
    ws_d_rank,
    "SEO Dashboard - 3-Week Ranking Distribution Breakdown",
    "Track movement across Top 1-3, Top 4-10, Striking Distance (11-20), Rank 21+, and Unindexed",
    max_col=8
)

w_labels = rank_3w.get("week_labels", ["W1", "W2", "W3"])

def add_rank_table(ws, start_r, title, rank_dict, total_cohort_count):
    ws.cell(row=start_r, column=1, value=title).font = section_font
    
    headers = [
        "Rank Bucket",
        f"W1 ({w_labels[0]}) Count", f"W1 %",
        f"W2 ({w_labels[1]}) Count", f"W2 %",
        f"W3 ({w_labels[2]}) Count", f"W3 %",
        "Net 3-Wk Change"
    ]
    style_header_row(ws, start_r + 1, headers)
    
    w1 = rank_dict.get("w1", {})
    w2 = rank_dict.get("w2", {})
    w3 = rank_dict.get("w3", {})
    
    buckets = [
        ("Top 1 - 3 (Rank 1-3)", "pos1_3"),
        ("Top 4 - 10 (Rank 4-10)", "pos4_10"),
        ("Striking Distance (Rank 11-20)", "pos11_20"),
        ("Rank 21+ (Beyond Top 20)", "pos21_plus"),
        ("Unindexed / > Rank 100", "unindexed")
    ]
    
    curr_r = start_r + 2
    for b_label, key in buckets:
        ws.row_dimensions[curr_r].height = 20
        
        v1 = w1.get(key, 0)
        v2 = w2.get(key, 0)
        v3 = w3.get(key, 0)
        diff = v3 - v1
        
        pct1 = (v1 / total_cohort_count) if total_cohort_count > 0 else 0
        pct2 = (v2 / total_cohort_count) if total_cohort_count > 0 else 0
        pct3 = (v3 / total_cohort_count) if total_cohort_count > 0 else 0
        
        c_lbl = ws.cell(row=curr_r, column=1, value=b_label)
        c_v1 = ws.cell(row=curr_r, column=2, value=v1)
        c_p1 = ws.cell(row=curr_r, column=3, value=pct1)
        c_v2 = ws.cell(row=curr_r, column=4, value=v2)
        c_p2 = ws.cell(row=curr_r, column=5, value=pct2)
        c_v3 = ws.cell(row=curr_r, column=6, value=v3)
        c_p3 = ws.cell(row=curr_r, column=7, value=pct3)
        c_df = ws.cell(row=curr_r, column=8, value=diff)
        
        for c in [c_v1, c_v2, c_v3, c_df]:
            c.number_format = "#,##0"
        for p in [c_p1, c_p2, c_p3]:
            p.number_format = "0.0%"
            
        for c_i, c_cell in enumerate([c_lbl, c_v1, c_p1, c_v2, c_p2, c_v3, c_p3, c_df], 1):
            c_cell.font = data_font
            c_cell.border = border_cell
            c_cell.alignment = align_left if c_i == 1 else align_right
            
        curr_r += 1
        
    # Total Row
    ws.row_dimensions[curr_r].height = 22
    tot_v1 = sum(w1.values())
    tot_v2 = sum(w2.values())
    tot_v3 = sum(w3.values())
    
    c_lbl = ws.cell(row=curr_r, column=1, value="Total Portfolio Cohort")
    c_v1 = ws.cell(row=curr_r, column=2, value=tot_v1)
    c_p1 = ws.cell(row=curr_r, column=3, value=1.0)
    c_v2 = ws.cell(row=curr_r, column=4, value=tot_v2)
    c_p2 = ws.cell(row=curr_r, column=5, value=1.0)
    c_v3 = ws.cell(row=curr_r, column=6, value=tot_v3)
    c_p3 = ws.cell(row=curr_r, column=7, value=1.0)
    c_df = ws.cell(row=curr_r, column=8, value=tot_v3 - tot_v1)
    
    for c in [c_v1, c_v2, c_v3, c_df]:
        c.number_format = "#,##0"
    for p in [c_p1, c_p2, c_p3]:
        p.number_format = "0.0%"
        
    for c_i, c_cell in enumerate([c_lbl, c_v1, c_p1, c_v2, c_p2, c_v3, c_p3, c_df], 1):
        c_cell.font = bold_data_font
        c_cell.fill = total_fill
        c_cell.border = border_total
        c_cell.alignment = align_left if c_i == 1 else align_right
        
    return curr_r + 2

r_pos = 4
r_pos = add_rank_table(ws_d_rank, r_pos, "1. Overall Ranking Distribution (All 1,564 Blogs)", rank_3w.get("overall", {}), 1564)
r_pos = add_rank_table(ws_d_rank, r_pos, "2. New Strategy (Post-July 16) Ranking Distribution (536 Blogs)", rank_3w.get("new_strategy", {}), 536)
r_pos = add_rank_table(ws_d_rank, r_pos, "3. Legacy Strategy (Pre-July 16) Ranking Distribution (1,028 Blogs)", rank_3w.get("legacy_strategy", {}), 1028)

autofit_columns(ws_d_rank, max_cols=8)

# ------------------------------------------------------------------------------
# Sheet 3: Master Blog Inventory
# ------------------------------------------------------------------------------
ws_d_master = wb_dash.create_sheet(title="Master Blog Inventory")

apply_title_banner(
    ws_d_master,
    "Master Blog Portfolio & SEO Metrics Export",
    "Complete dataset of 1,564 published WordPress blogs with Search Console and indexing metrics",
    max_col=14
)

master_headers = [
    "WP ID", "Blog Title", "URL", "Published Date", "Strategy Segment",
    "Indexing Status", "First Indexed Date", "Clicks (30d)", "Impressions (30d)", "CTR (%)",
    "Average Position", "W1 Rank", "W2 Rank", "W3 Rank"
]
style_header_row(ws_d_master, 4, master_headers)

for idx, b in enumerate(all_blogs, start=5):
    ws_d_master.row_dimensions[idx].height = 19
    
    is_ind = b.get("is_indexed", False)
    st_badge = "Indexed" if is_ind else "Pending Index"
    
    wr = b.get("weekly_rank", {})
    w1_p = wr.get("w1_pos", 0)
    w2_p = wr.get("w2_pos", 0)
    w3_p = wr.get("w3_pos", 0)
    
    c_id = ws_d_master.cell(row=idx, column=1, value=b.get("wp_id"))
    c_ttl = ws_d_master.cell(row=idx, column=2, value=b.get("title"))
    c_url = ws_d_master.cell(row=idx, column=3, value=b.get("link"))
    c_pub = ws_d_master.cell(row=idx, column=4, value=b.get("published_date"))
    c_str = ws_d_master.cell(row=idx, column=5, value="New Strategy" if b.get("is_new_strategy") else "Legacy Strategy")
    c_idx = ws_d_master.cell(row=idx, column=6, value=st_badge)
    c_fidx = ws_d_master.cell(row=idx, column=7, value=b.get("first_indexed_date", "Not Indexed Yet"))
    c_clk = ws_d_master.cell(row=idx, column=8, value=b.get("clicks", 0))
    c_imp = ws_d_master.cell(row=idx, column=9, value=b.get("impressions", 0))
    c_ctr = ws_d_master.cell(row=idx, column=10, value=b.get("ctr", 0) / 100.0)
    c_pos = ws_d_master.cell(row=idx, column=11, value=b.get("position", 0))
    c_w1 = ws_d_master.cell(row=idx, column=12, value=w1_p if w1_p > 0 else "Unindexed")
    c_w2 = ws_d_master.cell(row=idx, column=13, value=w2_p if w2_p > 0 else "Unindexed")
    c_w3 = ws_d_master.cell(row=idx, column=14, value=w3_p if w3_p > 0 else "Unindexed")
    
    c_id.number_format = "0"
    c_clk.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    c_pos.number_format = "0.0"
    
    for c_i, c_cell in enumerate([c_id, c_ttl, c_url, c_pub, c_str, c_idx, c_fidx, c_clk, c_imp, c_ctr, c_pos, c_w1, c_w2, c_w3], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        if c_i in (1, 4, 5, 6, 7):
            c_cell.alignment = align_center
        elif c_i in (2, 3):
            c_cell.alignment = align_left
        else:
            c_cell.alignment = align_right
            
    if is_ind:
        c_idx.fill = badge_indexed_fill
        c_idx.font = badge_indexed_font
    else:
        c_idx.fill = badge_pending_fill
        c_idx.font = badge_pending_font

autofit_columns(ws_d_master, max_cols=14)

# ------------------------------------------------------------------------------
# Sheet 4: New Strategy (Post-July 16)
# ------------------------------------------------------------------------------
ws_d_new = wb_dash.create_sheet(title="New Strategy (Post-July 16)")

apply_title_banner(
    ws_d_new,
    "🚀 New Strategy Blogs Export (Published >= July 16, 2026)",
    "Performance and Indexing audit for 536 New Strategy blogs",
    max_col=10
)

style_header_row(ws_d_new, 4, ["WP ID", "Blog Title", "URL", "Published Date", "Indexing Status", "First Indexed Date", "Clicks (30d)", "Impressions (30d)", "CTR (%)", "Average Position"])

for idx, b in enumerate(post_july16_blogs, start=5):
    ws_d_new.row_dimensions[idx].height = 19
    is_ind = b.get("is_indexed", False)
    st_badge = "Indexed" if is_ind else "Pending Index"
    
    c_id = ws_d_new.cell(row=idx, column=1, value=b.get("wp_id"))
    c_ttl = ws_d_new.cell(row=idx, column=2, value=b.get("title"))
    c_url = ws_d_new.cell(row=idx, column=3, value=b.get("link"))
    c_pub = ws_d_new.cell(row=idx, column=4, value=b.get("published_date"))
    c_idx = ws_d_new.cell(row=idx, column=5, value=st_badge)
    c_fidx = ws_d_new.cell(row=idx, column=6, value=b.get("first_indexed_date", "Not Indexed Yet"))
    c_clk = ws_d_new.cell(row=idx, column=7, value=b.get("clicks", 0))
    c_imp = ws_d_new.cell(row=idx, column=8, value=b.get("impressions", 0))
    c_ctr = ws_d_new.cell(row=idx, column=9, value=b.get("ctr", 0) / 100.0)
    c_pos = ws_d_new.cell(row=idx, column=10, value=b.get("position", 0))
    
    c_id.number_format = "0"
    c_clk.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    c_pos.number_format = "0.0"
    
    for c_i, c_cell in enumerate([c_id, c_ttl, c_url, c_pub, c_idx, c_fidx, c_clk, c_imp, c_ctr, c_pos], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        if c_i in (1, 4, 5, 6):
            c_cell.alignment = align_center
        elif c_i in (2, 3):
            c_cell.alignment = align_left
        else:
            c_cell.alignment = align_right
            
    if is_ind:
        c_idx.fill = badge_indexed_fill
        c_idx.font = badge_indexed_font
    else:
        c_idx.fill = badge_pending_fill
        c_idx.font = badge_pending_font

autofit_columns(ws_d_new, max_cols=10)

# ------------------------------------------------------------------------------
# Sheet 5: Legacy Strategy (Pre-July 16)
# ------------------------------------------------------------------------------
ws_d_leg = wb_dash.create_sheet(title="Legacy Strategy (Pre-July 16)")

apply_title_banner(
    ws_d_leg,
    "📜 Legacy Strategy Blogs Export (Published < July 16, 2026)",
    "Performance audit for 1,028 mature legacy blogs",
    max_col=10
)

style_header_row(ws_d_leg, 4, ["WP ID", "Blog Title", "URL", "Published Date", "Indexing Status", "First Indexed Date", "Clicks (30d)", "Impressions (30d)", "CTR (%)", "Average Position"])

for idx, b in enumerate(pre_july16_blogs, start=5):
    ws_d_leg.row_dimensions[idx].height = 19
    is_ind = b.get("is_indexed", False)
    st_badge = "Indexed" if is_ind else "Pending Index"
    
    c_id = ws_d_leg.cell(row=idx, column=1, value=b.get("wp_id"))
    c_ttl = ws_d_leg.cell(row=idx, column=2, value=b.get("title"))
    c_url = ws_d_leg.cell(row=idx, column=3, value=b.get("link"))
    c_pub = ws_d_leg.cell(row=idx, column=4, value=b.get("published_date"))
    c_idx = ws_d_leg.cell(row=idx, column=5, value=st_badge)
    c_fidx = ws_d_leg.cell(row=idx, column=6, value=b.get("first_indexed_date", "Mature Index"))
    c_clk = ws_d_leg.cell(row=idx, column=7, value=b.get("clicks", 0))
    c_imp = ws_d_leg.cell(row=idx, column=8, value=b.get("impressions", 0))
    c_ctr = ws_d_leg.cell(row=idx, column=9, value=b.get("ctr", 0) / 100.0)
    c_pos = ws_d_leg.cell(row=idx, column=10, value=b.get("position", 0))
    
    c_id.number_format = "0"
    c_clk.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    c_pos.number_format = "0.0"
    
    for c_i, c_cell in enumerate([c_id, c_ttl, c_url, c_pub, c_idx, c_fidx, c_clk, c_imp, c_ctr, c_pos], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        if c_i in (1, 4, 5, 6):
            c_cell.alignment = align_center
        elif c_i in (2, 3):
            c_cell.alignment = align_left
        else:
            c_cell.alignment = align_right
            
    if is_ind:
        c_idx.fill = badge_indexed_fill
        c_idx.font = badge_indexed_font
    else:
        c_idx.fill = badge_pending_fill
        c_idx.font = badge_pending_font

autofit_columns(ws_d_leg, max_cols=10)

# ------------------------------------------------------------------------------
# Sheet 6: Striking Distance Cohort
# ------------------------------------------------------------------------------
ws_d_strik = wb_dash.create_sheet(title="Striking Distance Cohort")

apply_title_banner(
    ws_d_strik,
    "🟡 Striking Distance Opportunity Cohort (Rank 11-20)",
    "57 High Potential Blogs ranking on Page 2 with massive impression growth potential",
    max_col=10
)

strik_headers = [
    "WP ID", "Blog Title", "URL", "Published Date", "Strategy Segment",
    "Clicks (30d)", "Impressions (30d)", "CTR (%)", "Average Position", "Primary Target Query"
]
style_header_row(ws_d_strik, 4, strik_headers)

for idx, b in enumerate(striking_blogs, start=5):
    ws_d_strik.row_dimensions[idx].height = 19
    
    top_q = b.get("top_queries", [])
    primary_q = top_q[0]["query"] if top_q else "N/A"
    
    c_id = ws_d_strik.cell(row=idx, column=1, value=b.get("wp_id"))
    c_ttl = ws_d_strik.cell(row=idx, column=2, value=b.get("title"))
    c_url = ws_d_strik.cell(row=idx, column=3, value=b.get("link"))
    c_pub = ws_d_strik.cell(row=idx, column=4, value=b.get("published_date"))
    c_str = ws_d_strik.cell(row=idx, column=5, value="New Strategy" if b.get("is_new_strategy") else "Legacy Strategy")
    c_clk = ws_d_strik.cell(row=idx, column=6, value=b.get("clicks", 0))
    c_imp = ws_d_strik.cell(row=idx, column=7, value=b.get("impressions", 0))
    c_ctr = ws_d_strik.cell(row=idx, column=8, value=b.get("ctr", 0) / 100.0)
    c_pos = ws_d_strik.cell(row=idx, column=9, value=b.get("position", 0))
    c_qry = ws_d_strik.cell(row=idx, column=10, value=primary_q)
    
    c_id.number_format = "0"
    c_clk.number_format = "#,##0"
    c_imp.number_format = "#,##0"
    c_ctr.number_format = "0.00%"
    c_pos.number_format = "0.0"
    
    for c_i, c_cell in enumerate([c_id, c_ttl, c_url, c_pub, c_str, c_clk, c_imp, c_ctr, c_pos, c_qry], 1):
        c_cell.font = data_font
        c_cell.border = border_cell
        if c_i in (1, 4, 5):
            c_cell.alignment = align_center
        elif c_i in (2, 3, 10):
            c_cell.alignment = align_left
        else:
            c_cell.alignment = align_right

autofit_columns(ws_d_strik, max_cols=10)

# Save Report 2
dash_path = BASE_DIR / "SEO_Dashboard_Report.xlsx"
wb_dash.save(dash_path)
print(f"✅ Generated {dash_path.name}")
